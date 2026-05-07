# Databricks notebook source
# MAGIC %md
# MAGIC # 01 — Bronze: SharePoint ingest
# MAGIC
# MAGIC Lists files in the mock SharePoint /input folder (a UC Volume on
# MAGIC Databricks; a local folder during dev) and registers each workbook in
# MAGIC `bronze.raw_workbooks`. Files themselves stay in the Volume — the
# MAGIC bronze table is metadata only.

# COMMAND ----------

dbutils.widgets.text("catalog", "quality_de")
dbutils.widgets.text("volume_input", "sharepoint_input")
CATALOG = dbutils.widgets.get("catalog")
VOL_IN = dbutils.widgets.get("volume_input")

# COMMAND ----------

import sys
sys.path.insert(0, "../")

from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

INPUT_PATH = Path(f"/Volumes/{CATALOG}/bronze/{VOL_IN}")
print(f"reading workbooks from {INPUT_PATH}")

# COMMAND ----------

rows = []
for p in sorted(INPUT_PATH.glob("*.xlsx")):
    wb = load_workbook(p, read_only=True, data_only=True)
    rows.append({
        "workbook": p.name,
        "source_path": str(p),
        "sheet_count": len(wb.sheetnames),
        "sheet_names": wb.sheetnames,
        "size_bytes": p.stat().st_size,
        "ingest_ts": datetime.utcnow(),
    })
    wb.close()
    print(f"  {p.name:<48} sheets={len(wb.sheetnames)}")

# COMMAND ----------

if rows:
    df = spark.createDataFrame(rows)
    (df.write.format("delta")
       .mode("overwrite")
       .option("overwriteSchema", "true")
       .saveAsTable(f"{CATALOG}.bronze.raw_workbooks"))
    display(spark.table(f"{CATALOG}.bronze.raw_workbooks"))
else:
    print("no workbooks found — run 00_setup.py first")
