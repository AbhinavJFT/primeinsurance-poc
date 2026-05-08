# Databricks notebook source
# MAGIC %md
# MAGIC # 05 — Demo walkthrough
# MAGIC
# MAGIC Narrated cells you can run live in front of an audience to show
# MAGIC before/after. No new logic — this notebook just visualizes the data
# MAGIC produced by 01–04.

# COMMAND ----------

dbutils.widgets.text("catalog", "quality_de")
CATALOG = dbutils.widgets.get("catalog")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 1. The mess — open one source sheet directly

# COMMAND ----------

import sys
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd

sys.path.insert(0, "../")

src = next(Path(f"/Volumes/{CATALOG}/bronze/sharepoint_input").glob("*.xlsx"))
wb = load_workbook(src, data_only=True)
ws = wb["4. D-2301"] if "4. D-2301" in wb.sheetnames else wb.worksheets[0]

print(f"raw sheet '{ws.title}' from {src.name}")
rows = list(ws.iter_rows(min_row=1, max_row=18, values_only=True))
display(pd.DataFrame(rows))

# COMMAND ----------

# MAGIC %md
# MAGIC ## 2. AI column mapping — what the model decided

# COMMAND ----------

display(spark.sql(f"""
SELECT sheet, raw_label, role, canonical, confidence, rationale, source
FROM {CATALOG}.silver.column_mapping_log
ORDER BY confidence ASC
LIMIT 30
"""))

# COMMAND ----------

# MAGIC %md
# MAGIC ## 3. DQ issues — every fix is logged

# COMMAND ----------

display(spark.sql(f"""
SELECT rule, severity, COUNT(*) AS issue_count
FROM {CATALOG}.silver.dq_issues
GROUP BY rule, severity
ORDER BY issue_count DESC
"""))

# COMMAND ----------

display(spark.sql(f"""
SELECT sheet, row_seq, column, rule, raw_value, repaired_value, note
FROM {CATALOG}.silver.dq_issues
WHERE severity = 'unparseable'
LIMIT 20
"""))

# COMMAND ----------

# MAGIC %md
# MAGIC ## 4. The clean output — long-form fact table

# COMMAND ----------

display(spark.sql(f"""
SELECT batch_no, sheet, analyte_canonical, value, spec_max, pass
FROM {CATALOG}.gold.fact_observation
WHERE pass = false
ORDER BY workbook, sheet, batch_no
LIMIT 20
"""))

# COMMAND ----------

# MAGIC %md
# MAGIC ## 5. Compliance summary

# COMMAND ----------

display(spark.sql(f"""
SELECT
  COUNT(*)                                       AS total_obs,
  SUM(CASE WHEN pass = true  THEN 1 ELSE 0 END)  AS passing,
  SUM(CASE WHEN pass = false THEN 1 ELSE 0 END)  AS failing,
  ROUND(AVG(CASE WHEN pass = true THEN 100.0
                 WHEN pass = false THEN 0.0 END), 2)  AS pass_rate_pct
FROM {CATALOG}.gold.fact_observation
"""))
