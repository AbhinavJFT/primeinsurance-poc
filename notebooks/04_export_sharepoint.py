# Databricks notebook source
# MAGIC %md
# MAGIC # 04 — Export cleaned workbooks to SharePoint /output
# MAGIC
# MAGIC Reads `gold.fact_observation` and writes one tidy single-sheet .xlsx per
# MAGIC source workbook to the mock SharePoint /output folder (a UC Volume on
# MAGIC Databricks; replace `SharePointMock` with a Microsoft Graph client to
# MAGIC point at real SharePoint).

# COMMAND ----------

dbutils.widgets.text("catalog", "quality_de")
dbutils.widgets.text("volume_output", "sharepoint_output")
CATALOG = dbutils.widgets.get("catalog")
VOL_OUT = dbutils.widgets.get("volume_output")

# COMMAND ----------

import shutil
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, "../")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path(f"/Volumes/{CATALOG}/bronze/{VOL_OUT}")
OUTPUT_PATH.mkdir(parents=True, exist_ok=True)
# UC Volumes are FUSE-mounted and don't support seek() during xlsx zip writes;
# build each workbook in a local temp dir, then stream-copy to the Volume.
_LOCAL_STAGING = Path(tempfile.mkdtemp(prefix="quality_export_"))

# COMMAND ----------

obs = spark.table(f"{CATALOG}.gold.fact_observation")
dq = spark.table(f"{CATALOG}.silver.dq_issues")
mapping = spark.table(f"{CATALOG}.silver.column_mapping_log")

workbooks = [r.workbook for r in obs.select("workbook").distinct().collect()]
print(f"writing cleaned workbooks for: {workbooks}")

# COMMAND ----------

def _style_header(ws, headers):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
        ws.column_dimensions[get_column_letter(j)].width = 18
    ws.freeze_panes = "A2"


for wb_name in workbooks:
    wb = Workbook()
    wb.remove(wb.active)

    obs_pdf = (obs.where(f"workbook = '{wb_name}'")
                  .orderBy("sheet", "row_seq", "analyte").toPandas())
    dq_pdf = dq.where(f"workbook = '{wb_name}'").toPandas()
    map_pdf = mapping.where(f"workbook = '{wb_name}'").toPandas()

    obs_ws = wb.create_sheet("observations")
    _style_header(obs_ws, list(obs_pdf.columns))
    for i, row in enumerate(obs_pdf.itertuples(index=False), start=2):
        for j, v in enumerate(row, start=1):
            obs_ws.cell(row=i, column=j, value=v)

    dq_ws = wb.create_sheet("dq_issues")
    _style_header(dq_ws, list(dq_pdf.columns))
    for i, row in enumerate(dq_pdf.itertuples(index=False), start=2):
        for j, v in enumerate(row, start=1):
            dq_ws.cell(row=i, column=j, value=v)

    map_ws = wb.create_sheet("column_mapping_log")
    _style_header(map_ws, list(map_pdf.columns))
    for i, row in enumerate(map_pdf.itertuples(index=False), start=2):
        for j, v in enumerate(row, start=1):
            map_ws.cell(row=i, column=j, value=v)

    staged = _LOCAL_STAGING / wb_name.replace(".xlsx", "_CLEAN.xlsx")
    wb.save(staged)
    target = OUTPUT_PATH / staged.name
    shutil.copy(staged, target)
    print(f"  wrote {target}  ({target.stat().st_size:,} bytes)")
