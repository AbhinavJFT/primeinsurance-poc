"""Activity handlers for the mock ADF runner.

Each handler takes ``(activity, ctx)`` and returns the activity's "output"
dict — the same shape the ``activity('X').output.*`` expression resolves
against in downstream steps.

Supported activity types: GetMetadata, ForEach, Copy, Lookup, Script.

Add new types here as pipelines need them. Keep handlers small — anything
non-trivial belongs in linked_services.py or a dedicated helper.
"""

from __future__ import annotations

import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .expressions import evaluate, evaluate_parameters, ExpressionContext
from .linked_services import (
    DeltaBackend,
    SharePointBackend,
    now_utc,
    resolve_backend,
)


# ---------------------------------------------------------------------------
# Runtime context (shared across activities of one pipeline run)
# ---------------------------------------------------------------------------

@dataclass
class RunContext:
    pipeline_parameters: dict[str, Any]
    activity_outputs: dict[str, Any]
    linked_services: dict[str, dict[str, Any]]
    datasets: dict[str, dict[str, Any]]
    spark: Any | None = None
    item: Any | None = None
    log: list[str] | None = None

    def expr_ctx(self, dataset_parameters: dict[str, Any] | None = None) -> ExpressionContext:
        return ExpressionContext(
            pipeline_parameters=self.pipeline_parameters,
            activity_outputs=self.activity_outputs,
            item=self.item,
            dataset_parameters=dataset_parameters,
        )

    def with_item(self, item: Any) -> "RunContext":
        return RunContext(
            pipeline_parameters=self.pipeline_parameters,
            activity_outputs=self.activity_outputs,
            linked_services=self.linked_services,
            datasets=self.datasets,
            spark=self.spark,
            item=item,
            log=self.log,
        )

    def emit(self, msg: str) -> None:
        if self.log is not None:
            self.log.append(msg)
        print(msg)


# ---------------------------------------------------------------------------
# Helpers — dataset / linked-service resolution at execution time
# ---------------------------------------------------------------------------

def _resolve_dataset_with_params(
    ds_ref: dict[str, Any], ctx: RunContext,
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Return (dataset_json, evaluated_parameters)."""
    ds_name = ds_ref["referenceName"]
    ds_json = ctx.datasets[ds_name]
    raw_params = ds_ref.get("parameters", {}) or {}
    params = evaluate_parameters(raw_params, ctx.expr_ctx())
    return ds_json, params


def _backend_for_dataset(ds_json: dict[str, Any], ctx: RunContext):
    ls_name = ds_json["properties"]["linkedServiceName"]["referenceName"]
    return resolve_backend(ctx.linked_services[ls_name], spark=ctx.spark)


# ---------------------------------------------------------------------------
# GetMetadata
# ---------------------------------------------------------------------------

def execute_get_metadata(activity: dict[str, Any], ctx: RunContext) -> dict[str, Any]:
    props = activity["typeProperties"]
    ds_json, ds_params = _resolve_dataset_with_params(props["dataset"], ctx)
    backend = _backend_for_dataset(ds_json, ctx)

    # Resolve folder path against dataset parameters.
    location = ds_json["properties"]["typeProperties"]["location"]
    folder = evaluate(location.get("folderPath", ""),
                      ctx.expr_ctx(ds_params))

    suffix = (props.get("filter") or {}).get("extension")

    if isinstance(backend, SharePointBackend):
        items = backend.list_files(folder, suffix=suffix)
    else:
        raise NotImplementedError(
            f"GetMetadata against {type(backend).__name__} is not wired in this POC"
        )

    ctx.emit(f"  [GetMetadata] folder={folder!r} found {len(items)} item(s)")
    return {"childItems": items}


# ---------------------------------------------------------------------------
# ForEach
# ---------------------------------------------------------------------------

def execute_for_each(activity: dict[str, Any], ctx: RunContext) -> dict[str, Any]:
    props = activity["typeProperties"]
    items = evaluate(props["items"], ctx.expr_ctx())
    items = list(items or [])
    inner = props.get("activities", [])

    ctx.emit(f"  [ForEach] {len(items)} iteration(s) over {len(inner)} inner activit(ies)")

    for item in items:
        item_ctx = ctx.with_item(item)
        # Local activity_outputs keep things readable for nested @activity() lookups,
        # but we share the parent's dict so successor activities outside the loop
        # can still reference the last iteration if they wish.
        for inner_activity in _topo_sort(inner):
            _dispatch(inner_activity, item_ctx)

    return {"iterations": len(items)}


# ---------------------------------------------------------------------------
# Copy
# ---------------------------------------------------------------------------

def execute_copy(activity: dict[str, Any], ctx: RunContext) -> dict[str, Any]:
    src_ref = activity["inputs"][0]
    dst_ref = activity["outputs"][0]

    src_ds, src_params = _resolve_dataset_with_params(src_ref, ctx)
    dst_ds, dst_params = _resolve_dataset_with_params(dst_ref, ctx)

    src_backend = _backend_for_dataset(src_ds, ctx)
    dst_backend = _backend_for_dataset(dst_ds, ctx)

    # Branch on backend pair — extend here if new combinations show up.
    if isinstance(src_backend, SharePointBackend) and isinstance(dst_backend, DeltaBackend):
        return _copy_sharepoint_to_volume(
            activity, ctx, src_backend, src_ds, src_params,
            dst_backend, dst_ds, dst_params,
        )

    if isinstance(src_backend, DeltaBackend) and isinstance(dst_backend, SharePointBackend):
        return _copy_delta_to_sharepoint_xlsx(
            activity, ctx, src_backend, src_ds, src_params,
            dst_backend, dst_ds, dst_params,
        )

    raise NotImplementedError(
        f"Copy not implemented for {type(src_backend).__name__} → {type(dst_backend).__name__}"
    )


def _copy_sharepoint_to_volume(
    activity, ctx, src_backend, src_ds, src_params, dst_backend, dst_ds, dst_params,
):
    src_loc = src_ds["properties"]["typeProperties"]["location"]
    folder = evaluate(src_loc["folderPath"], ctx.expr_ctx(src_params))
    file_name = src_params.get("fileName") or evaluate(
        src_loc.get("fileName", ""), ctx.expr_ctx(src_params)
    )

    payload = src_backend.read_bytes(folder, file_name)

    dst_loc = dst_ds["properties"]["typeProperties"]["location"]
    volume = dst_params.get("volume") or evaluate(
        dst_loc.get("folderPath", ""), ctx.expr_ctx(dst_params)
    )
    dst_name = dst_params.get("fileName") or evaluate(
        dst_loc.get("fileName", ""), ctx.expr_ctx(dst_params)
    )
    target = dst_backend.write_file(volume, dst_name, payload)

    ctx.emit(f"  [Copy] {folder}/{file_name}  →  {target}  ({len(payload):,} bytes)")
    return {
        "rowsCopied": 1,
        "bytesWritten": len(payload),
        "source": f"{folder}/{file_name}",
        "sink": str(target),
    }


def _copy_delta_to_sharepoint_xlsx(
    activity, ctx, src_backend, src_ds, src_params, dst_backend, dst_ds, dst_params,
):
    """Build a 3-sheet xlsx and upload to SharePoint /output. Honors a
    ``subfolder`` pipeline parameter (e.g. ``sessions/<sid>/transformed``)
    and a ``session_id`` parameter for filtering source rows."""
    workbook_name = src_params.get("workbook")
    if not workbook_name:
        raise ValueError("Delta→SharePoint copy requires dataset parameter 'workbook'")

    mock_cfg = src_ds["properties"].get("_mock") or {}
    companions = mock_cfg.get("companions") or {}

    primary = src_ds["properties"]["typeProperties"]
    primary_table = f"{primary['database']}.{primary['table']}"

    session_id = ctx.pipeline_parameters.get("session_id")
    fetch_session = None if session_id == "legacy_main_pipeline" else session_id

    sheets: dict[str, tuple[list[str], list[list[Any]]]] = {}
    sheets["observations"] = src_backend.fetch_for_workbook(
        primary_table, workbook_name, session_id=fetch_session,
    )
    for sheet_name, ref in companions.items():
        table = f"{ref['database']}.{ref['table']}"
        sheets[sheet_name] = src_backend.fetch_for_workbook(
            table, workbook_name, session_id=fetch_session,
        )

    out_loc = dst_ds["properties"]["typeProperties"]["location"]
    out_folder = evaluate(out_loc["folderPath"], ctx.expr_ctx(dst_params))
    out_name = dst_params.get("fileName") or evaluate(
        out_loc.get("fileName", ""), ctx.expr_ctx(dst_params)
    )

    # Optional pipeline-level subfolder (e.g. sessions/<sid>/transformed).
    # SharePointMock.upload_file accepts a slash-separated name and creates
    # subdirectories as needed.
    subfolder = ctx.pipeline_parameters.get("subfolder", "") or ""
    upload_name = f"{subfolder}/{out_name}" if subfolder else out_name

    with tempfile.TemporaryDirectory() as tmp:
        staged = Path(tmp) / out_name
        _build_xlsx(staged, sheets)
        uploaded = dst_backend.upload_file(out_folder, staged, name=upload_name)

    total_rows = sum(len(rows) for _, rows in sheets.values())
    ctx.emit(
        f"  [Copy] gold:{workbook_name}  →  {uploaded['path']}  "
        f"({total_rows} rows across {len(sheets)} sheet(s))"
    )
    return {
        "rowsCopied": total_rows,
        "bytesWritten": uploaded["size"],
        "source": primary_table,
        "sink": uploaded["path"],
    }


def _build_xlsx(path: Path, sheets: dict[str, tuple[list[str], list[list[Any]]]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, (headers, rows) in sheets.items():
        ws = wb.create_sheet(sheet_name)
        for j, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=j, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
            ws.column_dimensions[get_column_letter(j)].width = 18
        ws.freeze_panes = "A2"
        for i, row in enumerate(rows, start=2):
            for j, v in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=v)
    wb.save(path)


# ---------------------------------------------------------------------------
# Lookup
# ---------------------------------------------------------------------------

def execute_lookup(activity: dict[str, Any], ctx: RunContext) -> dict[str, Any]:
    props = activity["typeProperties"]
    ds_json, _ = _resolve_dataset_with_params(props["dataset"], ctx)
    backend = _backend_for_dataset(ds_json, ctx)

    if not isinstance(backend, DeltaBackend):
        raise NotImplementedError("Lookup is only wired against DeltaBackend in this POC")

    # The pipeline JSON expresses the query for documentation, but the mock
    # interprets the *intent*: distinct workbooks from the primary table.
    table = f"{ds_json['properties']['typeProperties']['database']}." \
            f"{ds_json['properties']['typeProperties']['table']}"
    session_id = ctx.pipeline_parameters.get("session_id")
    if session_id == "legacy_main_pipeline":
        session_id = None  # legacy path lists across all sessions
    rows = backend.lookup_distinct_workbooks(table, session_id=session_id)

    scope = f" (session={session_id})" if session_id else ""
    ctx.emit(f"  [Lookup] {table}{scope} → {len(rows)} distinct workbook(s)")
    return {"value": rows, "count": len(rows), "firstRow": rows[0] if rows else None}


# ---------------------------------------------------------------------------
# Script — mock-only, used for the manifest registration in pl_ingest
# ---------------------------------------------------------------------------

def execute_script(activity: dict[str, Any], ctx: RunContext) -> dict[str, Any]:
    mock = activity.get("_mock") or {}
    handler = mock.get("handler")

    if handler == "register_workbook_manifest":
        return _register_workbook_manifest(activity, ctx)

    if handler == "build_same_format_xlsx":
        return _build_same_format_xlsx(activity, ctx)

    raise NotImplementedError(
        f"Script activity {activity['name']!r} has no _mock.handler this runner understands"
    )


def _register_workbook_manifest(activity: dict[str, Any], ctx: RunContext) -> dict[str, Any]:
    """Stash the current item into a per-pipeline buffer.

    Called inside ForEach — we accumulate one row per workbook, then flush
    to bronze.raw_workbooks (Delta) or a JSON manifest (local) at the end of
    the pipeline. The flush happens in runner.run_pipeline after all
    activities complete.
    """
    item = ctx.item or {}
    name = item.get("name")
    size = item.get("size", 0)

    catalog = ctx.pipeline_parameters.get("catalog", "quality_de")
    volume = ctx.pipeline_parameters.get("volume", "sharepoint_input")
    session_id = ctx.pipeline_parameters.get("session_id", "legacy_main_pipeline")

    ls_db = next(
        ls for ls in ctx.linked_services.values()
        if ls["properties"]["type"] == "AzureDatabricksDeltaLake"
    )
    backend = resolve_backend(ls_db, spark=ctx.spark)
    if not isinstance(backend, DeltaBackend):
        raise RuntimeError("Delta backend resolution failed")

    landing = backend.landing_path(volume, name)
    row = {
        "workbook": name,
        "source_path": str(landing),
        "size_bytes": int(size),
        "ingest_ts": now_utc(),
        "catalog": catalog,
        "volume": volume,
        "session_id": session_id,
    }
    ctx.activity_outputs.setdefault("_manifest_buffer", []).append(row)
    return {"buffered": True, "workbook": name}


def _build_same_format_xlsx(activity: dict[str, Any], ctx: RunContext) -> dict[str, Any]:
    """Build a cleaned xlsx that mirrors the input shape and upload it to a
    subfolder under SharePoint /output (default: ``cleaned``).

    Reads:
      * the original input xlsx via SharePointBackend (input volume)
      * silver.observations_long + silver.column_mapping_log via DeltaBackend

    Writes:
      * a same-shape xlsx into <output volume>/<subfolder>/<workbook>.
    """
    from quality_core.inplace_cleaner import build_same_format_xlsx

    item = ctx.item or {}
    workbook_name = item.get("workbook")
    if not workbook_name:
        raise ValueError("build_same_format_xlsx requires ctx.item['workbook']")

    mock = activity.get("_mock") or {}
    input_folder = mock.get("inputFolder", "input")
    output_folder = mock.get("outputFolder", "output")
    subfolder = ctx.pipeline_parameters.get("subfolder") or mock.get("outputSubfolder", "cleaned")

    catalog = ctx.pipeline_parameters.get("catalog", "quality_de")

    # Resolve both backends from the linked services in the run context.
    sp_ls = next(
        ls for ls in ctx.linked_services.values()
        if ls["properties"]["type"] == "SharePointOnlineList"
    )
    db_ls = next(
        ls for ls in ctx.linked_services.values()
        if ls["properties"]["type"] == "AzureDatabricksDeltaLake"
    )
    sp_backend = resolve_backend(sp_ls)
    db_backend = resolve_backend(db_ls, spark=ctx.spark)
    if not isinstance(db_backend, DeltaBackend):
        raise RuntimeError("DeltaBackend resolution failed")

    session_id = ctx.pipeline_parameters.get("session_id")
    fetch_session = None if session_id == "legacy_main_pipeline" else session_id

    obs_headers, obs_rows = db_backend.fetch_for_workbook(
        f"silver.observations_long", workbook_name, session_id=fetch_session,
    )
    map_headers, map_rows = db_backend.fetch_for_workbook(
        f"silver.column_mapping_log", workbook_name, session_id=fetch_session,
    )

    # Stage the input + output xlsx in a tmp dir; then upload the cleaned
    # file to SharePoint /output/<subfolder>/.
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        input_staged = tmp_path / workbook_name
        input_staged.write_bytes(sp_backend.read_bytes(input_folder, workbook_name))
        output_staged = tmp_path / "cleaned" / workbook_name
        output_staged.parent.mkdir(parents=True, exist_ok=True)

        build_same_format_xlsx(
            input_path=input_staged,
            obs_headers=obs_headers,
            obs_rows=obs_rows,
            map_headers=map_headers,
            map_rows=map_rows,
            output_path=output_staged,
        )

        # Upload via SharePointMock.upload_file with name="cleaned/<file>"
        # so the subfolder is created on the fly.
        result = sp_backend.upload_file(
            output_folder, output_staged, name=f"{subfolder}/{workbook_name}",
        )

    ctx.emit(
        f"  [Script:same_format] {workbook_name} → {result['path']}  "
        f"({result['size']:,} bytes; {len(obs_rows)} obs rows)"
    )
    return {
        "rowsCopied": len(obs_rows),
        "bytesWritten": result["size"],
        "source": workbook_name,
        "sink": result["path"],
    }


# ---------------------------------------------------------------------------
# Dispatch + topological order (for ForEach inner activities)
# ---------------------------------------------------------------------------

_HANDLERS = {
    "GetMetadata": execute_get_metadata,
    "ForEach": execute_for_each,
    "Copy": execute_copy,
    "Lookup": execute_lookup,
    "Script": execute_script,
}


def dispatch(activity: dict[str, Any], ctx: RunContext) -> dict[str, Any]:
    return _dispatch(activity, ctx)


def _dispatch(activity: dict[str, Any], ctx: RunContext) -> dict[str, Any]:
    handler = _HANDLERS.get(activity["type"])
    if handler is None:
        raise ValueError(f"unsupported activity type: {activity['type']!r}")
    out = handler(activity, ctx)
    ctx.activity_outputs[activity["name"]] = {"output": out}
    return out


def _topo_sort(activities: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Order activities so each runs after all its dependsOn predecessors."""
    by_name = {a["name"]: a for a in activities}
    deps = {
        a["name"]: [d["activity"] for d in (a.get("dependsOn") or [])]
        for a in activities
    }

    ordered: list[dict[str, Any]] = []
    visited: set[str] = set()

    def visit(name: str, stack: tuple[str, ...] = ()) -> None:
        if name in visited:
            return
        if name in stack:
            raise ValueError(f"cycle in dependsOn through {stack + (name,)}")
        for d in deps.get(name, []):
            if d in by_name:
                visit(d, stack + (name,))
        visited.add(name)
        ordered.append(by_name[name])

    for a in activities:
        visit(a["name"])
    return ordered
