"""Build a same-format cleaned xlsx by overwriting dirty cells in a copy
of the original input.

The input xlsx is preserved exactly (merged cells, header band rows 3–11,
column widths, conditional formatting). For each data row we:

  1. Look up the cleaned meta values (date, time, batch, appearance, ...)
  2. Look up cleaned analyte values for that (sheet, row_seq, analyte_canonical)
  3. Replace only the data cells — header band is never touched.

Cells that the cleaner nulled (malformed time, negative impurity, NULL/?
sentinels) appear visibly empty in the output. The reason for each blanked
cell is in `silver.dq_issues`.
"""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# Header band rows that we never overwrite (matches generate_quality_data.py).
HEADER_BAND_LAST_ROW = 11
DATA_BAND_FIRST_ROW = 13   # row 12 is left blank as a divider in the source

# Map a canonical meta name to the column in the wide source layout.
# The mapping log tells us *which* raw column carries the canonical name; we
# look that up at runtime, so this dict is only used as a fallback ordering hint.
META_CANONICAL_FIELDS = (
    "sample_date",
    "sample_time",
    "report_time",
    "batch_no",
    "instrument_id",
    "stage",
    "sample_form",
    "appearance",
    "appearance_solution",
)


def _index_obs_rows(headers: list[str], rows: list[list[Any]]) -> dict[tuple[str, int], dict[str, Any]]:
    """(sheet, row_seq) → {meta dict + per-column-index value dict}.

    Each source row produces N output rows (one per analyte column). Meta
    fields are identical across the N, so we just take any of them. Analyte
    values are keyed by the source column index, so duplicate canonical
    names (e.g. three "Impurity-1" columns at different RTs) keep their
    distinct values.
    """
    h = {name: i for i, name in enumerate(headers)}
    out: dict[tuple[str, int], dict[str, Any]] = {}
    for r in rows:
        sheet = r[h["sheet"]]
        row_seq = r[h["row_seq"]]
        if row_seq is None:
            continue
        try:
            row_seq = int(row_seq)
        except (ValueError, TypeError):
            continue
        key = (sheet, row_seq)
        bucket = out.setdefault(key, {
            "meta": {f: r[h[f]] for f in META_CANONICAL_FIELDS if f in h},
            "value_by_column": {},
        })
        col_idx = r[h["column_index"]] if "column_index" in h else None
        value = r[h["value"]] if "value" in h else None
        if col_idx is not None:
            try:
                bucket["value_by_column"][int(col_idx)] = value
            except (ValueError, TypeError):
                pass
    return out


def _index_mappings(headers: list[str], rows: list[list[Any]]
                    ) -> dict[tuple[str, int], dict[str, Any]]:
    """(sheet, column_index) → {role, canonical}.

    Tells us which physical column in a sheet carries which canonical name.
    Used to put cleaned values back in the right cells.
    """
    h = {name: i for i, name in enumerate(headers)}
    out: dict[tuple[str, int], dict[str, Any]] = {}
    for r in rows:
        sheet = r[h["sheet"]]
        col_idx = r[h["column_index"]]
        if col_idx is None:
            continue
        try:
            col_idx = int(col_idx)
        except (ValueError, TypeError):
            continue
        out[(sheet, col_idx)] = {
            "role": r[h["role"]],
            "canonical": r[h["canonical"]],
        }
    return out


def _format_value_for_cell(value: Any) -> Any:
    """openpyxl writes None as empty cell, which is what we want for blanked
    cells. Times come through as datetime.time / datetime.datetime — keep them.
    Everything else is stringified by openpyxl correctly."""
    return value


def build_same_format_xlsx(
    input_path: Path | str,
    obs_headers: list[str],
    obs_rows: list[list[Any]],
    map_headers: list[str],
    map_rows: list[list[Any]],
    output_path: Path | str,
) -> Path:
    """Produce a cleaned xlsx that mirrors the input shape.

    Args:
        input_path:   the original messy xlsx (will be copied, not mutated)
        obs_headers:  column names of silver.observations_long
        obs_rows:     rows of silver.observations_long for this workbook only
        map_headers:  column names of silver.column_mapping_log
        map_rows:     rows of silver.column_mapping_log for this workbook only
        output_path:  destination path (parent must exist)
    """
    input_path = Path(input_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Copy the source to a working file so we keep all formatting intact;
    # we then only overwrite the data cells.
    shutil.copy2(input_path, output_path)

    obs_index = _index_obs_rows(obs_headers, obs_rows)
    map_index = _index_mappings(map_headers, map_rows)

    wb = load_workbook(output_path)
    cleaned_cells = 0
    blanked_cells = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Build a per-sheet column → mapping lookup.
        sheet_cols = {col_idx: info for (s, col_idx), info in map_index.items()
                      if s == sheet_name}
        if not sheet_cols:
            # MISC sheet (or anything we didn't profile) — leave untouched.
            continue

        # Walk the data band; rows below HEADER_BAND_LAST_ROW where col A
        # is numeric are the data rows.
        for r in range(DATA_BAND_FIRST_ROW, ws.max_row + 1):
            serial = ws.cell(row=r, column=1).value
            if not isinstance(serial, (int, float)):
                continue
            row_seq = int(serial)
            obs = obs_index.get((sheet_name, row_seq))
            if obs is None:
                continue

            for col_idx, info in sheet_cols.items():
                role = info["role"]
                canonical = info["canonical"]
                if not canonical:
                    continue

                if role == "meta":
                    cleaned = obs["meta"].get(canonical)
                elif role == "analyte":
                    # Look up by source column index to handle duplicate
                    # canonical names (e.g. multiple "Impurity-1" columns).
                    cleaned = obs["value_by_column"].get(col_idx)
                else:
                    continue

                cell = ws.cell(row=r, column=col_idx)
                # Don't change the cell if the cleaned value matches the
                # source — preserves number formatting and avoids spurious
                # writes (especially for values openpyxl reads back as float).
                if cell.value == cleaned:
                    continue
                cell.value = _format_value_for_cell(cleaned)
                if cleaned is None:
                    blanked_cells += 1
                else:
                    cleaned_cells += 1

    wb.save(output_path)
    return output_path


__all__ = ["build_same_format_xlsx"]
