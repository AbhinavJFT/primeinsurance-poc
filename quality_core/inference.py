"""Sheet schema inference.

Given a raw worksheet, locate the header band, identify meta columns
(date, batch, instrument, etc.) vs impurity columns, and return a
SheetProfile that downstream cleaning/mapping can consume.

Two layouts are supported:
  - quality_standard: the multi-row HPLC-style layout with RT/RRT/Spec
                      bands above the data (the layout in the screenshots).
  - misc_flat:        a flat single-row header layout (the MISC sheet).
"""

from __future__ import annotations

import re
from typing import Iterable

from openpyxl.worksheet.worksheet import Worksheet

from .models import HeaderBand, ImpurityProfile, MetaColumn, SheetProfile


# Anchor labels we look for in column A of the header band.
ANCHOR_LABELS = {
    "rt": "rt_row",
    "rrt": "rrt_row",
    "unit": "unit_row",
    "specification": "spec_bound_row",
    "internal srf specification": "internal_spec_row",
}

# Field labels we expect in the "field label row" of the standard layout.
EXPECTED_FIELD_LABELS = [
    "date of sampling", "time of sampling", "time of reporting",
    "batch no.", "batch no", "batch number",
    "instrument id", "stage", "sample characteristic", "sample characteristics",
    "appearance", "appearance of 30% solution",
]

# How many rows from the top to scan for headers (real sheets keep header
# bands within the first ~15 rows; we go a bit further for safety).
SCAN_LIMIT_ROWS = 25


def _normalize(s: object) -> str:
    if s is None:
        return ""
    return str(s).strip().lower()


def _row_values(ws: Worksheet, row_idx: int, max_col: int) -> list[object]:
    return [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]


def _looks_like_misc(ws: Worksheet) -> bool:
    """A flat sheet whose first non-empty row of strings has 4+ short labels."""
    max_col = min(ws.max_column, 12)
    for row_idx in range(1, 8):
        vals = _row_values(ws, row_idx, max_col)
        strs = [_normalize(v) for v in vals if v is not None]
        if len(strs) >= 4 and all(0 < len(s) <= 24 for s in strs):
            non_anchor = [s for s in strs if s not in ANCHOR_LABELS]
            if len(non_anchor) == len(strs):
                return True
    return False


def _detect_anchors(ws: Worksheet, max_col: int) -> dict[str, int]:
    """Find rows whose column A holds one of our anchor labels."""
    found: dict[str, int] = {}
    for row_idx in range(1, SCAN_LIMIT_ROWS + 1):
        label = _normalize(ws.cell(row=row_idx, column=1).value)
        if label in ANCHOR_LABELS:
            attr = ANCHOR_LABELS[label]
            found.setdefault(attr, row_idx)
    return found


def _detect_field_label_row(ws: Worksheet, max_col: int) -> int | None:
    """Find the row that has Date of Sampling / Batch No. / etc."""
    for row_idx in range(1, SCAN_LIMIT_ROWS + 1):
        labels = [_normalize(v) for v in _row_values(ws, row_idx, max_col)]
        joined = " | ".join(labels)
        hits = sum(1 for needle in EXPECTED_FIELD_LABELS if needle in joined)
        if hits >= 4:
            return row_idx
    return None


def _detect_process_label_row(ws: Worksheet, field_label_row: int) -> int | None:
    """Process label row sits 1–2 rows above field labels and contains 'inprocess' or 'analysis'."""
    for r in range(max(1, field_label_row - 3), field_label_row):
        joined = " | ".join(_normalize(v) for v in _row_values(ws, r, ws.max_column))
        if "inprocess" in joined or "analysis" in joined:
            return r
    return None


def _detect_impurity_columns(ws: Worksheet, rt_row: int, max_col: int) -> tuple[int, int]:
    """Return (first_col, last_col) for the impurity band — columns where RT row holds a number."""
    first = None
    last = None
    for c in range(1, max_col + 1):
        v = ws.cell(row=rt_row, column=c).value
        if isinstance(v, (int, float)):
            if first is None:
                first = c
            last = c
    if first is None or last is None:
        raise ValueError("Could not locate impurity columns from RT row")
    return first, last


def _safe_float(v: object) -> float | None:
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip())
        except ValueError:
            return None
    return None


def _safe_str(v: object) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _profile_quality_standard(ws: Worksheet) -> SheetProfile:
    max_col = ws.max_column

    anchors = _detect_anchors(ws, max_col)
    field_label_row = _detect_field_label_row(ws, max_col)
    if field_label_row is None or "rt_row" not in anchors:
        return SheetProfile(
            sheet_name=ws.title, layout="unknown", header=None,
            notes=["could not locate quality_standard header band"],
        )

    process_label_row = _detect_process_label_row(ws, field_label_row)
    rt_row = anchors["rt_row"]
    rrt_row = anchors.get("rrt_row")
    unit_row = anchors.get("unit_row")
    spec_bound_row = anchors.get("spec_bound_row")
    internal_spec_row = anchors.get("internal_spec_row")
    spec_value_row = (spec_bound_row + 1) if spec_bound_row else None

    first_imp_col, last_imp_col = _detect_impurity_columns(ws, rt_row, max_col)

    # Data start = first row after Internal SRF (+1 to skip blank divider) where col A is numeric.
    seek_from = (internal_spec_row or spec_value_row or rt_row) + 1
    data_start_row = seek_from
    for r in range(seek_from, ws.max_row + 1):
        if isinstance(ws.cell(row=r, column=1).value, (int, float)):
            data_start_row = r
            break

    header = HeaderBand(
        process_label_row=process_label_row,
        field_label_row=field_label_row,
        rt_row=rt_row,
        rrt_row=rrt_row,
        unit_row=unit_row,
        spec_bound_row=spec_bound_row,
        spec_value_row=spec_value_row,
        internal_spec_row=internal_spec_row,
        data_start_row=data_start_row,
    )

    # Meta columns: every column to the LEFT of the impurity band whose field-label
    # row cell carries a non-empty string (column A is the serial; we keep it but
    # mark its raw_label as "S.No" for downstream clarity).
    meta_columns: list[MetaColumn] = []
    for c in range(1, first_imp_col):
        raw = ws.cell(row=field_label_row, column=c).value
        label = _safe_str(raw) or ("S.No" if c == 1 else "")
        if label:
            meta_columns.append(MetaColumn(raw_label=label, column_index=c))

    # Impurity columns: pull RT/RRT/Unit/SpecBound/SpecValue/InternalSpec from the band.
    impurity_columns: list[ImpurityProfile] = []
    for c in range(first_imp_col, last_imp_col + 1):
        label = _safe_str(ws.cell(row=field_label_row, column=c).value) or f"col_{c}"
        impurity_columns.append(ImpurityProfile(
            raw_label=label,
            column_index=c,
            rt=_safe_float(ws.cell(row=rt_row, column=c).value),
            rrt=_safe_float(ws.cell(row=rrt_row, column=c).value) if rrt_row else None,
            unit=_safe_str(ws.cell(row=unit_row, column=c).value) if unit_row else None,
            sub_unit=_safe_str(ws.cell(row=unit_row + 1, column=c).value) if unit_row else None,
            spec_bound=_safe_str(ws.cell(row=spec_bound_row, column=c).value) if spec_bound_row else None,
            spec_value=_safe_float(ws.cell(row=spec_value_row, column=c).value) if spec_value_row else None,
            internal_spec_value=_safe_float(ws.cell(row=internal_spec_row, column=c).value) if internal_spec_row else None,
        ))

    # Count data rows (rows where col A is numeric, from data_start_row onward).
    n_data = 0
    for r in range(data_start_row, ws.max_row + 1):
        if isinstance(ws.cell(row=r, column=1).value, (int, float)):
            n_data += 1

    return SheetProfile(
        sheet_name=ws.title,
        layout="quality_standard",
        header=header,
        meta_columns=meta_columns,
        impurity_columns=impurity_columns,
        n_data_rows=n_data,
    )


def _profile_misc_flat(ws: Worksheet) -> SheetProfile:
    max_col = ws.max_column
    field_label_row = None
    for row_idx in range(1, 8):
        vals = _row_values(ws, row_idx, max_col)
        strs = [_normalize(v) for v in vals if v is not None]
        if len(strs) >= 4 and all(0 < len(s) <= 24 for s in strs):
            field_label_row = row_idx
            break

    if field_label_row is None:
        return SheetProfile(sheet_name=ws.title, layout="unknown", header=None,
                            notes=["misc_flat detector tripped but header row not isolated"])

    meta_columns = []
    for c in range(1, max_col + 1):
        v = ws.cell(row=field_label_row, column=c).value
        label = _safe_str(v)
        if label:
            meta_columns.append(MetaColumn(raw_label=label, column_index=c))

    data_start_row = field_label_row + 1
    while data_start_row <= ws.max_row and all(
        ws.cell(row=data_start_row, column=mc.column_index).value in (None, "")
        for mc in meta_columns
    ):
        data_start_row += 1

    n_data = 0
    for r in range(data_start_row, ws.max_row + 1):
        if any(ws.cell(row=r, column=mc.column_index).value not in (None, "")
               for mc in meta_columns):
            n_data += 1

    header = HeaderBand(
        process_label_row=None,
        field_label_row=field_label_row,
        rt_row=None, rrt_row=None, unit_row=None,
        spec_bound_row=None, spec_value_row=None, internal_spec_row=None,
        data_start_row=data_start_row,
    )

    return SheetProfile(
        sheet_name=ws.title,
        layout="misc_flat",
        header=header,
        meta_columns=meta_columns,
        n_data_rows=n_data,
    )


def profile_sheet(ws: Worksheet) -> SheetProfile:
    """Inspect the worksheet and return a structured SheetProfile."""
    if ws.max_row < 3 or ws.max_column < 2:
        return SheetProfile(sheet_name=ws.title, layout="unknown", header=None,
                            notes=["sheet too small to profile"])

    max_col = ws.max_column
    anchors = _detect_anchors(ws, max_col)
    if "rt_row" in anchors and "spec_bound_row" in anchors:
        return _profile_quality_standard(ws)
    if _looks_like_misc(ws):
        return _profile_misc_flat(ws)
    return SheetProfile(sheet_name=ws.title, layout="unknown", header=None,
                        notes=["no anchors and no flat header detected"])
