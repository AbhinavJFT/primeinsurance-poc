"""Orchestrate inference -> mapping -> cleaning -> pivot for a workbook."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

from .cleaning import (
    coerce_appearance,
    coerce_date,
    coerce_float,
    coerce_pass_fail,
    coerce_str,
    coerce_time,
    evaluate_pass,
)
from .inference import profile_sheet
from .mapping import load_schema, map_columns
from .models import ColumnMapping, DQIssue, ImpurityProfile, Observation, SheetProfile


@dataclass
class WorkbookCleanResult:
    workbook: str
    profiles: list[SheetProfile] = field(default_factory=list)
    mappings: list[ColumnMapping] = field(default_factory=list)
    observations: list[Observation] = field(default_factory=list)
    misc_rows: list[dict[str, Any]] = field(default_factory=list)
    dq_issues: list[DQIssue] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_META_COERCERS = {
    "sample_date": coerce_date,
    "sample_time": coerce_time,
    "report_time": coerce_time,
    "batch_no": coerce_str,
    "instrument_id": coerce_str,
    "stage": coerce_str,
    "sample_form": coerce_str,
    "appearance": coerce_appearance,
    "appearance_solution": coerce_pass_fail,
}


def _spec_bounds(imp: ImpurityProfile) -> tuple[Optional[float], Optional[float],
                                                Optional[float], Optional[float]]:
    """Return (spec_min, spec_max, internal_min, internal_max)."""
    if imp.spec_bound and imp.spec_bound.lower().startswith("min"):
        return imp.spec_value, None, imp.internal_spec_value, None
    return None, imp.spec_value, None, imp.internal_spec_value


# ---------------------------------------------------------------------------
# Sheet processors
# ---------------------------------------------------------------------------

def _process_quality_sheet(ws, profile: SheetProfile, mappings: list[ColumnMapping],
                           workbook_name: str, result: WorkbookCleanResult) -> None:
    header = profile.header
    assert header is not None
    meta_map: dict[int, ColumnMapping] = {m.column_index: m for m in mappings if m.role == "meta"}
    analyte_map: dict[int, ColumnMapping] = {m.column_index: m for m in mappings
                                             if m.role == "analyte"}

    for r in range(header.data_start_row, ws.max_row + 1):
        serial = ws.cell(row=r, column=1).value
        if not isinstance(serial, (int, float)):
            continue
        row_seq = int(serial)

        # Coerce meta values once per row.
        meta_clean: dict[str, Any] = {}
        for col_idx, mapping in meta_map.items():
            raw = ws.cell(row=r, column=col_idx).value
            coercer = _META_COERCERS.get(mapping.canonical or "")
            if coercer is None:
                value, rule = coerce_str(raw)
            else:
                value, rule = coercer(raw)
            meta_clean[mapping.canonical] = value
            if rule:
                result.dq_issues.append(DQIssue(
                    workbook=workbook_name, sheet=profile.sheet_name,
                    row_seq=row_seq, column=mapping.canonical,
                    rule=rule, severity="unparseable" if rule.startswith("malformed") else "repaired",
                    raw_value=raw, repaired_value=value,
                    note=f"raw label was {mapping.raw_label!r}",
                ))

        # One Observation per impurity column in this row.
        for imp in profile.impurity_columns:
            mapping = analyte_map.get(imp.column_index)
            if mapping is None:
                continue
            raw = ws.cell(row=r, column=imp.column_index).value
            value, rule = coerce_float(raw)
            if rule:
                result.dq_issues.append(DQIssue(
                    workbook=workbook_name, sheet=profile.sheet_name,
                    row_seq=row_seq, column=imp.raw_label,
                    rule=rule,
                    severity="unparseable" if rule in ("negative_value", "non_numeric") else "repaired",
                    raw_value=raw, repaired_value=value,
                    note=f"analyte={mapping.canonical}",
                ))
            spec_min, spec_max, isrf_min, isrf_max = _spec_bounds(imp)
            passed = evaluate_pass(value, spec_min, spec_max)
            result.observations.append(Observation(
                workbook=workbook_name,
                sheet=profile.sheet_name,
                row_seq=row_seq,
                sample_date=meta_clean.get("sample_date"),
                sample_time=meta_clean.get("sample_time"),
                report_time=meta_clean.get("report_time"),
                batch_no=meta_clean.get("batch_no"),
                instrument_id=meta_clean.get("instrument_id"),
                stage=meta_clean.get("stage"),
                sample_form=meta_clean.get("sample_form"),
                appearance=meta_clean.get("appearance"),
                appearance_solution=("OK" if meta_clean.get("appearance_solution") is True
                                     else "FAIL" if meta_clean.get("appearance_solution") is False
                                     else None),
                analyte=imp.raw_label,
                analyte_canonical=mapping.canonical,
                column_index=imp.column_index,
                rt=imp.rt,
                rrt=imp.rrt,
                value=value,
                unit=(imp.unit or "") + ("/" + imp.sub_unit if imp.sub_unit else "") or None,
                spec_min=spec_min,
                spec_max=spec_max,
                spec_internal_min=isrf_min,
                spec_internal_max=isrf_max,
                pass_=passed,
                raw_value=None if raw is None else str(raw),
                mapping_confidence=mapping.confidence,
            ))


def _process_misc_sheet(ws, profile: SheetProfile, mappings: list[ColumnMapping],
                        workbook_name: str, result: WorkbookCleanResult) -> None:
    header = profile.header
    assert header is not None
    cols = {m.column_index: m.raw_label for m in mappings}
    for r in range(header.data_start_row, ws.max_row + 1):
        row: dict[str, Any] = {"workbook": workbook_name, "sheet": profile.sheet_name,
                               "row_seq": r - header.data_start_row + 1}
        non_empty = False
        for col_idx, label in cols.items():
            raw = ws.cell(row=r, column=col_idx).value
            if raw not in (None, ""):
                non_empty = True
            cleaned, rule = coerce_str(raw)
            row[label] = cleaned
            if rule:
                result.dq_issues.append(DQIssue(
                    workbook=workbook_name, sheet=profile.sheet_name,
                    row_seq=row["row_seq"], column=label, rule=rule,
                    severity="repaired", raw_value=raw, repaired_value=cleaned,
                ))
        if non_empty:
            result.misc_rows.append(row)


# ---------------------------------------------------------------------------
# Top-level entry point
# ---------------------------------------------------------------------------

def process_workbook(path: Path | str, *, schema: Optional[dict] = None,
                     llm_client=None) -> WorkbookCleanResult:
    path = Path(path)
    schema = schema or load_schema()
    wb = load_workbook(path, data_only=True)
    workbook_name = path.name
    result = WorkbookCleanResult(workbook=workbook_name)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        profile = profile_sheet(ws)
        result.profiles.append(profile)

        if profile.layout == "unknown":
            continue

        mappings = map_columns(profile, schema, workbook=workbook_name,
                               llm_client=llm_client)
        result.mappings.extend(mappings)

        if profile.layout == "quality_standard":
            _process_quality_sheet(ws, profile, mappings, workbook_name, result)
        elif profile.layout == "misc_flat":
            _process_misc_sheet(ws, profile, mappings, workbook_name, result)

    return result


# ---------------------------------------------------------------------------
# Tidy xlsx writer
# ---------------------------------------------------------------------------

_OBS_HEADERS = [
    "workbook", "sheet", "row_seq", "sample_date", "sample_time", "report_time",
    "batch_no", "instrument_id", "stage", "sample_form", "appearance",
    "appearance_solution", "analyte", "analyte_canonical", "rt", "rrt",
    "value", "unit", "spec_min", "spec_max", "spec_internal_min",
    "spec_internal_max", "pass", "raw_value", "mapping_confidence",
]
_DQ_HEADERS = ["workbook", "sheet", "row_seq", "column", "rule",
               "severity", "raw_value", "repaired_value", "note"]
_MAP_HEADERS = ["workbook", "sheet", "column_index", "raw_label", "role",
                "canonical", "confidence", "rationale", "source"]


def _write_table(ws, headers: list[str], rows: list[list[Any]]) -> None:
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    for i, row in enumerate(rows, start=2):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)
    ws.freeze_panes = "A2"
    for col_idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18


def write_tidy_workbook(result: WorkbookCleanResult, path: Path | str) -> Path:
    """Write a single cleaned workbook with sheets: observations, dq_issues, column_mapping, misc."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    obs_ws = wb.create_sheet("observations")
    obs_rows: list[list[Any]] = []
    for o in result.observations:
        obs_rows.append([
            o.workbook, o.sheet, o.row_seq, o.sample_date, o.sample_time, o.report_time,
            o.batch_no, o.instrument_id, o.stage, o.sample_form, o.appearance,
            o.appearance_solution, o.analyte, o.analyte_canonical, o.rt, o.rrt,
            o.value, o.unit, o.spec_min, o.spec_max, o.spec_internal_min,
            o.spec_internal_max, o.pass_, o.raw_value, o.mapping_confidence,
        ])
    _write_table(obs_ws, _OBS_HEADERS, obs_rows)

    dq_ws = wb.create_sheet("dq_issues")
    dq_rows = [[d.workbook, d.sheet, d.row_seq, d.column, d.rule, d.severity,
                str(d.raw_value), str(d.repaired_value), d.note]
               for d in result.dq_issues]
    _write_table(dq_ws, _DQ_HEADERS, dq_rows)

    map_ws = wb.create_sheet("column_mapping_log")
    map_rows = [[m.workbook, m.sheet, m.column_index, m.raw_label, m.role,
                 m.canonical, m.confidence, m.rationale, m.source]
                for m in result.mappings]
    _write_table(map_ws, _MAP_HEADERS, map_rows)

    if result.misc_rows:
        misc_ws = wb.create_sheet("misc")
        misc_headers = sorted({k for row in result.misc_rows for k in row.keys()})
        # Keep workbook/sheet/row_seq leading
        for lead in ["row_seq", "sheet", "workbook"]:
            if lead in misc_headers:
                misc_headers.remove(lead)
                misc_headers.insert(0, lead)
        misc_rows = [[row.get(h) for h in misc_headers] for row in result.misc_rows]
        _write_table(misc_ws, misc_headers, misc_rows)

    wb.save(path)
    return path
