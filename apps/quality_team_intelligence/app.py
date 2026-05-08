"""Quality Team Intelligence — interactive Streamlit app.

Three top-level views via the sidebar:

  Home      — upload an .xlsx (or generate the 3 demo files), preview,
              click Clean. Live-streaming progress, then a per-run
              results page with Deliverables / DQ Audit / Column Resolution /
              Compliance Metrics tabs (all scoped to the file just processed).

  History   — every previous interactive run, newest first. Click any
              row to re-open that run's results.

  Dashboard — comprehensive analytics across both the interactive
              runs AND the batch pipeline's bronze/silver/gold tables.

Each interactive run is persisted to /Volumes/<catalog>/bronze/<output>/
_app_runs/<run_id>/ as a small directory of summary.json + parquet files
+ the input/output xlsx files. No SQL writes, so the app needs only
READ_VOLUME + WRITE_VOLUME on the output volume.
"""

from __future__ import annotations

import io
import json
import os
import shutil
import sys
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Vendored deps live alongside this file (see vendor.sh). When running
# locally from the repo root we also fall back to the repo's quality_core.
APP_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(APP_DIR))
if (APP_DIR.parent.parent / "quality_core").exists():
    sys.path.insert(0, str(APP_DIR.parent.parent))

from quality_core import process_workbook, write_tidy_workbook  # noqa: E402
from quality_core.inplace_cleaner import build_same_format_xlsx  # noqa: E402
from generate_quality_data import generate as generate_synthetic  # noqa: E402

# Try optional Databricks imports — only needed for SQL/job features
try:
    from databricks import sql as dbsql
    from databricks.sdk import WorkspaceClient
    from databricks.sdk.core import Config
    DBSDK_AVAILABLE = True
except Exception:
    DBSDK_AVAILABLE = False


# ===========================================================================
# Configuration
# ===========================================================================

CATALOG = os.environ.get("QDE_CATALOG", "quality_de")
VOLUME_OUTPUT = os.environ.get("QDE_VOLUME_OUTPUT", "sharepoint_output")
WAREHOUSE_ID = os.environ.get("DATABRICKS_WAREHOUSE_ID", "2de6a251cf2870eb")
JOB_NAME_HINT = os.environ.get("QDE_JOB_NAME_HINT", "quality_de")

# State storage. Databricks Apps don't FUSE-mount UC Volumes — volume access
# requires the Files API, which would mean refactoring all our pandas reads
# and writes. /tmp lives inside the app container, survives within a single
# app session, and is wiped on container restart. That's fine for a demo.
RUNS_DIR = Path(os.environ.get("QDE_RUNS_DIR", "/tmp/qde_runs"))
WORKING_DIR = Path(os.environ.get("QDE_WORKING_DIR", "/tmp/qde_working"))

# Visual constants
PRIMARY = "#2563EB"
ACCENT_GREEN = "#059669"
ACCENT_RED = "#DC2626"
ACCENT_AMBER = "#D97706"
NEUTRAL = "#475569"


# ===========================================================================
# Page setup
# ===========================================================================

st.set_page_config(
    page_title="Quality Team Intelligence",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
    <style>
        .block-container { padding-top: 3.5rem; padding-bottom: 2rem; max-width: 1400px; }
        h1 { font-size: 2.0rem; line-height: 1.25; margin-top: 0.2rem; }
        h2 { font-size: 1.4rem; margin-top: 0.5rem; }
        h3 { font-size: 1.1rem; opacity: 0.85; }

        [data-testid="stMetricValue"] { font-size: 1.7rem; font-weight: 700; }
        [data-testid="stMetricLabel"] { font-size: 0.78rem; opacity: 0.7;
                                         text-transform: uppercase; letter-spacing: 0.04em; }

        .stTabs [data-baseweb="tab-list"] { gap: 1.4rem; border-bottom: 1px solid #E2E8F0; }
        .stTabs [data-baseweb="tab"] { font-size: 0.95rem; padding: 0.55rem 0.2rem; }
        .stTabs [aria-selected="true"] { color: #2563EB !important; }

        .badge {
            display: inline-block;
            padding: 0.18rem 0.6rem;
            border-radius: 999px;
            font-size: 0.75rem;
            font-weight: 600;
            font-family: ui-monospace, SFMono-Regular, monospace;
        }
        .badge-ok    { background: #DCFCE7; color: #166534; }
        .badge-fail  { background: #FEE2E2; color: #991B1B; }
        .badge-warn  { background: #FEF3C7; color: #92400E; }
        .badge-info  { background: #DBEAFE; color: #1D4ED8; }
        .badge-mute  { background: #F1F5F9; color: #475569; }

        .card {
            border: 1px solid #E2E8F0;
            border-radius: 10px;
            padding: 1rem 1.2rem;
            background: #FFFFFF;
            transition: border-color .15s, box-shadow .15s;
        }
        .card:hover { border-color: #93C5FD; box-shadow: 0 1px 3px rgba(37, 99, 235, 0.08); }

        .runheader {
            background: linear-gradient(90deg, #EFF6FF 0%, #FAFAFA 100%);
            border: 1px solid #DBEAFE;
            border-radius: 10px;
            padding: 1rem 1.2rem;
            margin-bottom: 1.2rem;
        }

        .footer-note { opacity: 0.6; font-size: 0.85rem; padding-top: 1rem; }

        section[data-testid="stSidebar"] {
            background: linear-gradient(180deg, #F8FAFC 0%, #F1F5F9 100%);
            border-right: 1px solid #E2E8F0;
        }
        section[data-testid="stSidebar"] .stButton button {
            text-align: left;
            font-weight: 500;
            background: transparent;
            border: 1px solid transparent;
            box-shadow: none;
            color: #1E293B;
        }
        section[data-testid="stSidebar"] .stButton button:hover {
            background: rgba(37, 99, 235, 0.08);
            border-color: rgba(37, 99, 235, 0.2);
        }
    </style>
    """,
    unsafe_allow_html=True,
)


# ===========================================================================
# Session state
# ===========================================================================

def _init_state():
    defaults = {
        "view":              "home",
        "stage":             "empty",     # empty | staged | running | results
        "session_id":        None,        # current session_id (str)
        "session_n_files":   0,           # files generated this session
        "session_files":     [],          # list[str] file names in this session
        "run_id":             None,        # int — Databricks job run id
        "run_started_at":    None,        # float — time.time() when run was triggered
        "log_lines":         [],
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


_init_state()


# ===========================================================================
# Helpers — Databricks SDK + SQL
# ===========================================================================

@st.cache_resource(show_spinner=False)
def _config():
    if not DBSDK_AVAILABLE:
        return None
    try:
        return Config()
    except Exception:
        return None


@st.cache_resource(show_spinner=False)
def _ws_client():
    if not DBSDK_AVAILABLE:
        return None
    try:
        return WorkspaceClient()
    except Exception:
        return None


def _sql_conn():
    cfg = _config()
    if cfg is None:
        return None
    host = (cfg.host or os.environ.get("DATABRICKS_HOST", "")
            ).replace("https://", "").replace("http://", "")
    return dbsql.connect(
        server_hostname=host,
        http_path=f"/sql/1.0/warehouses/{WAREHOUSE_ID}",
        credentials_provider=lambda: cfg.authenticate,
    )


@st.cache_data(ttl=30, show_spinner=False)
def run_query(query: str) -> pd.DataFrame:
    conn = _sql_conn()
    if conn is None:
        return pd.DataFrame()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(query)
                cols = [c[0] for c in cur.description]
                rows = cur.fetchall()
        return pd.DataFrame(rows, columns=cols)
    except Exception as e:
        st.error(f"SQL query failed: `{e}`")
        return pd.DataFrame()


# ===========================================================================
# Helpers — file management
# ===========================================================================

def _ensure_dirs():
    """Create _app_runs/ and _app_working/ inside the output volume if they
    don't exist. Idempotent."""
    try:
        RUNS_DIR.mkdir(parents=True, exist_ok=True)
        WORKING_DIR.mkdir(parents=True, exist_ok=True)
    except (PermissionError, OSError):
        # Volume might not be writable — degrade gracefully
        pass


def _save_uploaded(uploaded_file) -> Path:
    """Save a Streamlit-uploaded file into the working dir."""
    _ensure_dirs()
    target = WORKING_DIR / uploaded_file.name
    target.write_bytes(uploaded_file.getbuffer())
    return target


def _generate_demo_files() -> list[Path]:
    """Run the synthetic generator → 3 standard demo workbooks."""
    _ensure_dirs()
    # Wipe working dir so we don't accumulate stale files
    if WORKING_DIR.exists():
        for p in WORKING_DIR.glob("*.xlsx"):
            try:
                p.unlink()
            except Exception:
                pass
    return generate_synthetic(WORKING_DIR, seed=43)


def _list_available_files() -> list[Path]:
    if not WORKING_DIR.exists():
        return []
    return sorted(WORKING_DIR.glob("*.xlsx"))


def _list_sheets(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _read_sheet_preview(path: Path, sheet_name: str, max_rows: int = 22) -> pd.DataFrame:
    """Compact tabular preview (used in dropdown previews)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append(row)
    finally:
        wb.close()
    if not rows:
        return pd.DataFrame()
    width = max(len(r) for r in rows)
    cols = [chr(65 + c) if c < 26 else f"col_{c+1}" for c in range(width)]
    padded = [list(r) + [None] * (width - len(r)) for r in rows]
    return pd.DataFrame(padded, columns=cols)


# ===========================================================================
# Helpers — Databricks session orchestration (session-scoped pipeline)
# ===========================================================================

INPUT_VOLUME_BASE = f"/Volumes/{CATALOG}/bronze/sharepoint_input/sessions"
OUTPUT_VOLUME_BASE = f"/Volumes/{CATALOG}/bronze/sharepoint_output/sessions"
PIPELINE_JOB_NAME_HINT = JOB_NAME_HINT  # bundle name "quality_de" matches the job


def _mint_session_id() -> str:
    """YYYY-MM-DD-hhmmss-<6hex> — human-readable, sortable, unique-enough."""
    import secrets
    now = datetime.now(timezone.utc)
    suffix = secrets.token_hex(3)
    return f"{now.strftime('%Y-%m-%d-%H%M%S')}-{suffix}"


def _resolve_pipeline_job_id() -> int:
    """Look up the medallion pipeline job's ID by name hint."""
    w = _ws_client()
    if w is None:
        raise RuntimeError("Workspace client unavailable; cannot trigger job")
    for j in w.jobs.list():
        name = (j.settings.name if j.settings else "") or ""
        if PIPELINE_JOB_NAME_HINT in name:
            return j.job_id
    raise RuntimeError(
        f"No job found whose name contains {PIPELINE_JOB_NAME_HINT!r}"
    )


def _generate_session_files(session_id: str, n_files: int) -> list[Path]:
    """Generate N synthetic workbooks for this session in /tmp/qde_session/<sid>/.
    Files alternate API/KSM/Intermediates with varied seeds so each looks
    distinct."""
    from generate_quality_data import build_workbook, workbook_specs
    local_dir = Path("/tmp/qde_session") / session_id
    local_dir.mkdir(parents=True, exist_ok=True)
    for old in local_dir.glob("*.xlsx"):
        try:
            old.unlink()
        except Exception:
            pass
    specs = workbook_specs()
    keys = list(specs.keys())
    written: list[Path] = []
    for i in range(n_files):
        spec = specs[keys[i % len(keys)]]
        wb = build_workbook(spec, seed=43 + i)
        base = spec.filename.replace(".xlsx", "")
        target = local_dir / f"{base}_{i:03d}.xlsx"
        wb.save(target)
        written.append(target)
    return written


def _upload_session_files(session_id: str, local_files: list[Path]) -> list[str]:
    """Upload local files to /Volumes/.../sharepoint_input/sessions/<sid>/.
    Returns the list of remote volume paths."""
    w = _ws_client()
    if w is None:
        raise RuntimeError("Workspace client unavailable; cannot upload")
    target_dir = f"{INPUT_VOLUME_BASE}/{session_id}"
    uploaded: list[str] = []
    for f in local_files:
        remote = f"{target_dir}/{f.name}"
        with open(f, "rb") as fh:
            w.files.upload(remote, fh, overwrite=True)
        uploaded.append(remote)
    return uploaded


def _trigger_pipeline(session_id: str) -> int:
    """Trigger the medallion job with this session_id. Returns run_id."""
    w = _ws_client()
    if w is None:
        raise RuntimeError("Workspace client unavailable; cannot trigger job")
    job_id = _resolve_pipeline_job_id()
    run = w.jobs.run_now(
        job_id=job_id,
        notebook_params={"session_id": session_id},
    )
    return run.run_id


def _poll_run(run_id: int) -> dict:
    """Single poll: returns dict with overall + per-task status."""
    w = _ws_client()
    if w is None:
        raise RuntimeError("Workspace client unavailable; cannot poll run")
    run = w.jobs.get_run(run_id=run_id)
    state = run.state
    return {
        "life_cycle_state": (state.life_cycle_state.value
                              if state and state.life_cycle_state else "UNKNOWN"),
        "result_state": (state.result_state.value
                          if state and state.result_state else None),
        "state_message": (state.state_message if state else "") or "",
        "tasks": [
            {
                "task_key": t.task_key,
                "life_cycle_state": (t.state.life_cycle_state.value
                                      if t.state and t.state.life_cycle_state else "PENDING"),
                "result_state": (t.state.result_state.value
                                  if t.state and t.state.result_state else None),
                "start_time": t.start_time,
                "end_time": t.end_time,
                "state_message": (t.state.state_message if t.state else "") or "",
            }
            for t in (run.tasks or [])
        ],
    }


def _clear_session_table_rows(session_id: str) -> None:
    """Delete rows tagged with session_id from all 5 tables. Used by Retry
    (re-trigger after partial run) and full Discard."""
    for table in [
        f"{CATALOG}.bronze.raw_workbooks",
        f"{CATALOG}.silver.observations_long",
        f"{CATALOG}.silver.dq_issues",
        f"{CATALOG}.silver.column_mapping_log",
        f"{CATALOG}.gold.fact_observation",
    ]:
        try:
            run_query(f"DELETE FROM {table} WHERE session_id = '{session_id}'")
        except Exception as e:
            print(f"  ({table} cleanup skipped: {e})")


def _discard_session(session_id: str) -> None:
    """Full discard: delete input subfolder AND any rows already written."""
    w = _ws_client()
    target_dir = f"{INPUT_VOLUME_BASE}/{session_id}"
    if w is not None:
        try:
            for f in w.files.list_directory_contents(target_dir):
                w.files.delete(f.path)
            w.files.delete_directory(target_dir)
        except Exception as e:
            print(f"  (input cleanup skipped: {e})")
    _clear_session_table_rows(session_id)


# ---------------------------------------------------------------------------
# Excel-shaped renderer (preserves merged cells, fills, fonts)
# ---------------------------------------------------------------------------

def _format_cell_value(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        if v.hour == 0 and v.minute == 0 and v.second == 0:
            return v.strftime("%Y-%m-%d")
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        s = f"{v:.4f}".rstrip("0").rstrip(".")
        return s if s else "0"
    return str(v)


def _argb_to_hex(color) -> str | None:
    """openpyxl colors come back as ARGB hex strings. Strip alpha; skip
    pure black/white/empty so we don't repaint default cells."""
    if color is None:
        return None
    s = color if isinstance(color, str) else getattr(color, "rgb", None)
    if not isinstance(s, str):
        return None
    s = s.upper()
    if len(s) == 8:
        s = s[2:]
    if s in ("000000", "FFFFFF", ""):
        return None
    if len(s) != 6:
        return None
    return f"#{s}"


def _xlsx_sheet_to_html(path: Path, sheet_name: str) -> str:
    """Render a sheet as an HTML <table> preserving merged cells, fills,
    fonts, alignment. Returns just the table markup (no <html> wrapper)."""
    wb = load_workbook(path, data_only=True)  # not read_only — we need merged_cells
    try:
        ws = wb[sheet_name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        # Merged ranges → top-left cell + skip set for the rest.
        merged_top: dict[tuple[int, int], tuple[int, int]] = {}
        skip_cells: set[tuple[int, int]] = set()
        for mrange in ws.merged_cells.ranges:
            r0, c0 = mrange.min_row, mrange.min_col
            r1, c1 = mrange.max_row, mrange.max_col
            merged_top[(r0, c0)] = (r1 - r0 + 1, c1 - c0 + 1)
            for rr in range(r0, r1 + 1):
                for cc in range(c0, c1 + 1):
                    if (rr, cc) != (r0, c0):
                        skip_cells.add((rr, cc))

        # Column widths in pixels (openpyxl exposes character widths).
        col_widths: dict[int, int] = {}
        for letter, dim in (ws.column_dimensions or {}).items():
            try:
                idx = ws[f"{letter}1"].column
                if dim.width:
                    col_widths[idx] = max(60, int(dim.width * 7))
            except Exception:
                pass

        out = ['<table class="xlsx">']

        # Top header row: A, B, C, …
        out.append('<thead><tr><th class="rowhdr"></th>')
        for c in range(1, max_col + 1):
            w = col_widths.get(c, 90)
            out.append(
                f'<th class="colhdr" style="min-width:{w}px;width:{w}px;">'
                f'{get_column_letter(c)}</th>'
            )
        out.append('</tr></thead><tbody>')

        for r in range(1, max_row + 1):
            out.append(f'<tr><td class="rowhdr">{r}</td>')
            for c in range(1, max_col + 1):
                if (r, c) in skip_cells:
                    continue
                cell = ws.cell(row=r, column=c)

                attrs: list[str] = []
                if (r, c) in merged_top:
                    rs, cs = merged_top[(r, c)]
                    if rs > 1:
                        attrs.append(f'rowspan="{rs}"')
                    if cs > 1:
                        attrs.append(f'colspan="{cs}"')

                styles: list[str] = []
                if cell.fill and cell.fill.fill_type == "solid":
                    bg = _argb_to_hex(cell.fill.start_color)
                    if bg:
                        styles.append(f"background-color:{bg}")
                if cell.font:
                    if cell.font.bold:
                        styles.append("font-weight:600")
                    if cell.font.italic:
                        styles.append("font-style:italic")
                    fc = _argb_to_hex(cell.font.color)
                    if fc:
                        styles.append(f"color:{fc}")
                if cell.alignment:
                    if cell.alignment.horizontal:
                        styles.append(f"text-align:{cell.alignment.horizontal}")
                    if cell.alignment.wrap_text:
                        styles.append("white-space:normal")

                value = cell.value
                if (isinstance(value, (int, float))
                        and not any("text-align" in s for s in styles)):
                    styles.append("text-align:right")

                style_attr = f'style="{";".join(styles)}"' if styles else ""
                attr_str = " ".join(attrs)
                display = (_format_cell_value(value)
                           .replace("&", "&amp;")
                           .replace("<", "&lt;")
                           .replace(">", "&gt;"))
                out.append(f"<td {attr_str} {style_attr}>{display}</td>")
            out.append('</tr>')

        out.append('</tbody></table>')
        return "\n".join(out)
    finally:
        wb.close()


def _render_xlsx_full(path: Path, sheet_name: str, height: int = 560):
    """Render the sheet inside an iframe so we can fully control styling
    (sticky row/col headers, merged-cell display) without fighting Streamlit's
    own CSS."""
    table_html = _xlsx_sheet_to_html(path, sheet_name)
    page = f"""
    <!doctype html>
    <html><head><meta charset="utf-8">
    <style>
        body {{ margin: 0; padding: 0;
                 font-family: -apple-system, BlinkMacSystemFont, "Segoe UI",
                              Roboto, Helvetica, Arial, sans-serif;
                 background: #FFFFFF; }}
        .scroller {{ overflow: auto; max-height: {height - 6}px;
                      border: 1px solid #CBD5E1; border-radius: 4px; }}
        table.xlsx {{ border-collapse: collapse; font-size: 12px; color: #0F172A; }}
        table.xlsx th, table.xlsx td {{
            border: 1px solid #E2E8F0;
            padding: 4px 6px;
            vertical-align: middle;
            background: #FFFFFF;
            white-space: nowrap;
        }}
        table.xlsx td {{ min-width: 70px; }}
        table.xlsx .colhdr {{
            background: #F1F5F9; color: #475569; font-weight: 600;
            text-align: center;
            position: sticky; top: 0; z-index: 2;
            border-bottom: 1px solid #94A3B8;
        }}
        table.xlsx .rowhdr {{
            background: #F1F5F9; color: #475569; font-weight: 500;
            text-align: center; min-width: 36px;
            position: sticky; left: 0; z-index: 1;
            border-right: 1px solid #94A3B8;
        }}
        table.xlsx thead .rowhdr {{ z-index: 3; }}
    </style></head>
    <body>
        <div class="scroller">{table_html}</div>
    </body></html>
    """
    components.html(page, height=height, scrolling=False)


# ===========================================================================
# Helpers — run persistence
# ===========================================================================

def _serialize_observations(observations) -> pd.DataFrame:
    return pd.DataFrame([
        {
            "workbook": o.workbook, "sheet": o.sheet, "row_seq": o.row_seq,
            "sample_date": o.sample_date,
            "sample_time": str(o.sample_time) if o.sample_time else None,
            "report_time": str(o.report_time) if o.report_time else None,
            "batch_no": o.batch_no, "instrument_id": o.instrument_id,
            "stage": o.stage, "sample_form": o.sample_form,
            "appearance": o.appearance, "appearance_solution": o.appearance_solution,
            "analyte": o.analyte, "analyte_canonical": o.analyte_canonical,
            "column_index": o.column_index,
            "rt": o.rt, "rrt": o.rrt, "value": o.value, "unit": o.unit,
            "spec_min": o.spec_min, "spec_max": o.spec_max,
            "spec_internal_min": o.spec_internal_min,
            "spec_internal_max": o.spec_internal_max,
            "pass": o.pass_, "raw_value": o.raw_value,
            "mapping_confidence": o.mapping_confidence,
        }
        for o in observations
    ])


def _serialize_dq_issues(dq_issues) -> pd.DataFrame:
    return pd.DataFrame([
        {
            "workbook": d.workbook, "sheet": d.sheet, "row_seq": d.row_seq,
            "column": d.column, "rule": d.rule, "severity": d.severity,
            "raw_value": str(d.raw_value), "repaired_value": str(d.repaired_value),
            "note": d.note,
        }
        for d in dq_issues
    ])


def _serialize_mappings(mappings) -> pd.DataFrame:
    return pd.DataFrame([
        {
            "workbook": m.workbook, "sheet": m.sheet,
            "column_index": m.column_index, "raw_label": m.raw_label,
            "role": m.role, "canonical": m.canonical,
            "confidence": m.confidence, "rationale": m.rationale,
            "source": m.source,
        }
        for m in mappings
    ])


def _save_run(
    run_id: str,
    input_file: Path,
    result,
    tidy_xlsx: Path,
    same_fmt_xlsx: Path,
    duration_s: float,
) -> Path:
    """Persist all artifacts for one run into RUNS_DIR/<run_id>/."""
    run_dir = RUNS_DIR / run_id
    run_dir.mkdir(parents=True, exist_ok=True)

    # Copy input + outputs
    shutil.copy(input_file, run_dir / "input.xlsx")
    shutil.copy(tidy_xlsx, run_dir / "tidy.xlsx")
    shutil.copy(same_fmt_xlsx, run_dir / "same_format.xlsx")

    # Serialize tables
    obs_df = _serialize_observations(result.observations)
    dq_df = _serialize_dq_issues(result.dq_issues)
    map_df = _serialize_mappings(result.mappings)

    obs_df.to_parquet(run_dir / "observations.parquet")
    dq_df.to_parquet(run_dir / "dq_issues.parquet")
    map_df.to_parquet(run_dir / "mappings.parquet")

    # Summary
    n_violations = int((obs_df["pass"] == False).sum()) if not obs_df.empty else 0
    n_pass = int((obs_df["pass"] == True).sum()) if not obs_df.empty else 0
    pass_rate = round(100 * n_pass / max(n_pass + n_violations, 1), 1)

    summary = {
        "run_id": run_id,
        "file_name": input_file.name,
        "input_size_bytes": int(input_file.stat().st_size),
        "started_at": datetime.now(timezone.utc).isoformat(),
        "duration_s": round(duration_s, 2),
        "n_observations": int(len(obs_df)),
        "n_dq_issues": int(len(dq_df)),
        "n_mappings": int(len(map_df)),
        "n_violations": n_violations,
        "n_pass": n_pass,
        "pass_rate_pct": pass_rate,
        "low_confidence_mappings": int(((map_df["confidence"] < 0.5)
                                        & (map_df["role"] != "ignored")).sum())
                                   if not map_df.empty else 0,
        "sheets_count": result.profiles and len(result.profiles) or 0,
    }
    (run_dir / "summary.json").write_text(json.dumps(summary, indent=2, default=str))
    return run_dir


def _list_runs() -> list[dict]:
    if not RUNS_DIR.exists():
        return []
    runs = []
    for d in sorted(RUNS_DIR.iterdir(), key=lambda p: p.name, reverse=True):
        if not d.is_dir():
            continue
        summary_path = d / "summary.json"
        if not summary_path.exists():
            continue
        try:
            summary = json.loads(summary_path.read_text())
            runs.append(summary)
        except Exception:
            continue
    return runs


def _load_run(run_id: str) -> dict | None:
    run_dir = RUNS_DIR / run_id
    summary_path = run_dir / "summary.json"
    if not summary_path.exists():
        return None
    summary = json.loads(summary_path.read_text())
    obs = pd.read_parquet(run_dir / "observations.parquet")
    dq = pd.read_parquet(run_dir / "dq_issues.parquet")
    mp = pd.read_parquet(run_dir / "mappings.parquet")
    return {
        "run_id": run_id,
        "dir": run_dir,
        "summary": summary,
        "observations": obs,
        "dq_issues": dq,
        "mappings": mp,
    }


# ===========================================================================
# Sidebar
# ===========================================================================

def _sidebar():
    with st.sidebar:
        st.markdown(
            "<div style='font-size:1.3rem;font-weight:700;'>Quality Team</div>"
            "<div style='font-size:1.0rem;font-weight:600;opacity:0.7;'>Intelligence</div>",
            unsafe_allow_html=True,
        )
        st.markdown("---")

        nav_items = [
            ("Home", "home"),
            ("History", "history"),
            ("Dashboard", "dashboard"),
        ]
        for label, key in nav_items:
            if st.button(label, key=f"nav_{key}", use_container_width=True):
                st.session_state.view = key
                if key == "home":
                    # Reset stage when going home, unless we're showing a result
                    if st.session_state.stage == "results":
                        pass  # keep results visible
                st.rerun()

        st.markdown("---")
        st.caption(f"Catalog: `{CATALOG}`")
        runs = _list_runs()
        st.caption(f"Stored runs: **{len(runs)}**")

        st.markdown(
            "<div class='footer-note' style='padding-top:1rem;'>"
            "Self-contained ADF mock<br>SharePoint round-trip<br>AI-driven cleaning"
            "</div>",
            unsafe_allow_html=True,
        )


# ===========================================================================
# Home view — upload / generate / preview / clean / results
# ===========================================================================

def _render_home_empty():
    st.title("Process a Quality team workbook")
    st.markdown(
        "<div style='opacity:0.75;font-size:1.05rem;margin-bottom:1.5rem;'>"
        "Generate synthetic Quality team workbooks and run them through the "
        "Databricks medallion pipeline. Each session runs in isolation — its "
        "files live in a session subfolder, and its results are tagged with a "
        "session ID so they don't mix with other runs."
        "</div>",
        unsafe_allow_html=True,
    )

    with st.container(border=True):
        st.markdown("### Generate session files")
        st.markdown(
            "<div style='opacity:0.65;font-size:0.9rem;margin-bottom:0.6rem;'>"
            "How many synthetic workbooks should this session produce? "
            "Files alternate between API, KSM, and Intermediates types. "
            "Cap is 25 to stay under the Foundation Model rate limit."
            "</div>",
            unsafe_allow_html=True,
        )
        n_files = st.slider(
            "Number of files",
            min_value=1, max_value=25, value=20, step=1,
            key="gen_count",
        )
        if st.button("Generate", type="primary",
                     use_container_width=True, key="btn_generate"):
            session_id = _mint_session_id()
            try:
                with st.spinner(f"Generating {n_files} workbooks…"):
                    local_files = _generate_session_files(session_id, n_files)
                with st.spinner(f"Uploading {n_files} workbooks to volume…"):
                    _upload_session_files(session_id, local_files)
            except Exception as e:
                st.error(f"Generate/upload failed: {e}")
                return
            st.session_state.session_id = session_id
            st.session_state.session_n_files = n_files
            st.session_state.session_files = [f.name for f in local_files]
            st.session_state.stage = "staged"
            st.toast(f"Generated {n_files} workbooks")
            st.rerun()


def _render_home_staged():
    st.title("Session ready")
    sid = st.session_state.get("session_id")
    n = st.session_state.get("session_n_files", 0)
    files = st.session_state.get("session_files", [])

    if not sid:
        st.session_state.stage = "empty"
        st.rerun()

    st.markdown(
        f"Session **`{sid}`** has **{n} file(s)** uploaded to "
        f"`{INPUT_VOLUME_BASE}/{sid}/`. Click **Run pipeline** to trigger "
        "the medallion job, or **Discard session** to clean up and start over."
    )

    with st.expander(f"Files in this session ({n})", expanded=False):
        for fn in files:
            st.markdown(f"- `{fn}`")

    c1, c2 = st.columns([3, 1])
    with c1:
        if st.button("Run pipeline", type="primary",
                     use_container_width=True, key="btn_run"):
            try:
                with st.spinner("Triggering pipeline…"):
                    run_id = _trigger_pipeline(sid)
            except Exception as e:
                st.error(f"Pipeline trigger failed: {e}")
                return
            st.session_state.run_id = run_id
            st.session_state.run_started_at = time.time()
            st.session_state.stage = "running"
            st.rerun()
    with c2:
        if st.button("Discard session", use_container_width=True, key="btn_discard"):
            with st.spinner("Discarding session…"):
                try:
                    _discard_session(sid)
                except Exception as e:
                    st.warning(f"Cleanup partial: {e}")
            for k in ("session_id", "session_n_files", "session_files",
                      "run_id", "run_started_at"):
                st.session_state.pop(k, None)
            st.session_state.stage = "empty"
            st.rerun()


def _render_home_running():
    st.title("Pipeline running")
    sid = st.session_state.get("session_id")
    run_id = st.session_state.get("run_id")
    started_at = st.session_state.get("run_started_at") or time.time()

    if not (sid and run_id):
        st.error("No active run.")
        st.session_state.stage = "empty"
        st.rerun()

    header = st.empty()
    body = st.empty()
    failure = st.empty()

    POLL_INTERVAL_S = 3
    TASK_ORDER = ["setup", "bronze_ingest", "silver_ai_cleaning",
                  "gold_curated", "export_sharepoint"]

    while True:
        elapsed = int(time.time() - started_at)
        try:
            poll = _poll_run(run_id)
        except Exception as e:
            header.error(f"Polling failed: {e}")
            time.sleep(POLL_INTERVAL_S)
            continue

        header.markdown(
            f"Session **`{sid}`** &nbsp;·&nbsp; pipeline run `{run_id}` "
            f"&nbsp;·&nbsp; **{elapsed // 60}m {elapsed % 60}s** elapsed"
        )

        tasks_by_key = {t["task_key"]: t for t in poll["tasks"]}
        rows = []
        for tk in TASK_ORDER:
            t = tasks_by_key.get(tk, {"life_cycle_state": "PENDING",
                                       "result_state": None,
                                       "start_time": None,
                                       "end_time": None})
            lcs = t.get("life_cycle_state", "PENDING")
            rs = t.get("result_state")
            if lcs == "TERMINATED" and rs == "SUCCESS":
                marker, status_text = "✓", "SUCCESS"
            elif lcs == "TERMINATED" and rs in ("FAILED", "TIMEDOUT", "CANCELED"):
                marker, status_text = "✗", rs
            elif lcs == "RUNNING":
                marker, status_text = "●", "RUNNING"
            elif lcs == "PENDING":
                marker, status_text = "○", "PENDING"
            elif lcs == "SKIPPED":
                marker, status_text = "—", "SKIPPED"
            else:
                marker, status_text = "○", lcs

            if t.get("start_time"):
                start_ms = t["start_time"]
                end_ms = t.get("end_time") or int(time.time() * 1000)
                t_secs = max(0, (end_ms - start_ms) // 1000)
                t_elapsed = (f"{t_secs // 60}m {t_secs % 60}s"
                             if t_secs >= 60 else f"{t_secs}s")
            else:
                t_elapsed = "—"
            rows.append(
                f"`{marker}` &nbsp; **{tk}** &nbsp;·&nbsp; "
                f"{status_text} &nbsp;·&nbsp; {t_elapsed}"
            )

        body.markdown("\n\n".join(rows))

        lcs_overall = poll["life_cycle_state"]
        if lcs_overall == "TERMINATED":
            rs_overall = poll["result_state"]
            if rs_overall == "SUCCESS":
                st.session_state.stage = "results"
                time.sleep(0.4)
                st.rerun()
            else:
                failed_tasks = [
                    t["task_key"] for t in poll["tasks"]
                    if t.get("result_state") in ("FAILED", "TIMEDOUT", "CANCELED")
                ]
                msg = poll.get("state_message") or "Pipeline did not complete successfully."
                failure.error(
                    f"Pipeline {rs_overall}. "
                    f"Failed task(s): {', '.join(failed_tasks) or 'unknown'}.\n\n{msg}"
                )
                cR, cD = st.columns(2)
                with cR:
                    if st.button("Retry this session", type="primary",
                                 use_container_width=True, key="btn_retry"):
                        try:
                            with st.spinner("Cleaning prior rows + re-triggering…"):
                                _clear_session_table_rows(sid)
                                run_id2 = _trigger_pipeline(sid)
                            st.session_state.run_id = run_id2
                            st.session_state.run_started_at = time.time()
                            st.rerun()
                        except Exception as e:
                            st.error(f"Retry failed: {e}")
                with cD:
                    if st.button("Discard and start over",
                                 use_container_width=True, key="btn_fail_discard"):
                        try:
                            with st.spinner("Discarding session…"):
                                _discard_session(sid)
                        except Exception as e:
                            st.warning(f"Cleanup partial: {e}")
                        for k in ("session_id", "session_n_files",
                                  "session_files", "run_id", "run_started_at"):
                            st.session_state.pop(k, None)
                        st.session_state.stage = "empty"
                        st.rerun()
                return
        elif lcs_overall == "INTERNAL_ERROR":
            failure.error(
                f"Pipeline INTERNAL_ERROR: {poll.get('state_message','')}"
            )
            if st.button("Back to staged", key="btn_back_to_staged"):
                st.session_state.stage = "staged"
                st.rerun()
            return
        time.sleep(POLL_INTERVAL_S)


def _load_session_observations(session_id: str) -> pd.DataFrame:
    return run_query(
        f"SELECT * FROM {CATALOG}.gold.fact_observation "
        f"WHERE session_id = '{session_id}'"
    )


def _load_session_dq_issues(session_id: str) -> pd.DataFrame:
    return run_query(
        f"SELECT * FROM {CATALOG}.silver.dq_issues "
        f"WHERE session_id = '{session_id}'"
    )


def _load_session_mappings(session_id: str) -> pd.DataFrame:
    return run_query(
        f"SELECT * FROM {CATALOG}.silver.column_mapping_log "
        f"WHERE session_id = '{session_id}'"
    )


def _render_home_results():
    sid = st.session_state.get("session_id")
    if not sid:
        st.error("No session in state.")
        st.session_state.stage = "empty"
        st.rerun()

    obs = _load_session_observations(sid)
    dq = _load_session_dq_issues(sid)
    mp = _load_session_mappings(sid)

    a1, a2 = st.columns([4, 1])
    with a1:
        st.title("Results")
        st.caption(
            f"Session **`{sid}`** &nbsp;·&nbsp; "
            f"{len(obs):,} observations &nbsp;·&nbsp; "
            f"{len(dq):,} DQ fixes &nbsp;·&nbsp; "
            f"{len(mp):,} mapping decisions"
        )
    with a2:
        if st.button("Run another", use_container_width=True, key="run_another"):
            for k in ("session_id", "session_n_files", "session_files",
                      "run_id", "run_started_at"):
                st.session_state.pop(k, None)
            st.session_state.stage = "empty"
            st.rerun()

    # KPI tiles
    if not obs.empty:
        n_obs = len(obs)
        n_pass = int((obs["pass"] == True).sum()) if "pass" in obs.columns else 0
        n_fail = int((obs["pass"] == False).sum()) if "pass" in obs.columns else 0
        decided = max(n_pass + n_fail, 1)
        pass_pct = round(100.0 * n_pass / decided, 1)
        n_workbooks = obs["workbook"].nunique() if "workbook" in obs.columns else 0

        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric("Workbooks", f"{n_workbooks:,}")
        k2.metric("Observations", f"{n_obs:,}")
        k3.metric("DQ fixes", f"{len(dq):,}")
        k4.metric("Spec violations", f"{n_fail:,}")
        k5.metric("Pass rate", f"{pass_pct}%")

    # Tabs (same shape as before, session-scoped data)
    out_clean = f"{OUTPUT_VOLUME_BASE}/{sid}/cleaned"
    out_tidy = f"{OUTPUT_VOLUME_BASE}/{sid}/transformed"

    tab_outputs, tab_clean, tab_ai, tab_analytics = st.tabs([
        "Deliverables", "DQ Audit", "Column Resolution", "Compliance Metrics",
    ])
    with tab_outputs:
        _tab_outputs_session(out_clean, out_tidy)
    with tab_clean:
        _tab_cleaning(dq)
    with tab_ai:
        _tab_ai_mapping(mp)
    with tab_analytics:
        _tab_analytics(obs)


def _tab_cleaning(dq: pd.DataFrame):
    st.markdown(
        "**Every cleaning rule that fired during this run.** "
        "Each repair is logged with raw → repaired so any reviewer can audit."
    )
    if dq.empty:
        st.info("No DQ events recorded.")
        return

    breakdown = dq.groupby(["rule", "severity"]).size().reset_index(name="issues")
    breakdown = breakdown.sort_values("issues", ascending=False)

    chart_col, table_col = st.columns([3, 2])
    with chart_col:
        st.markdown("**Fixes by rule**")
        st.bar_chart(breakdown.set_index("rule")["issues"], height=320, color=PRIMARY)
    with table_col:
        st.markdown("**Total**")
        st.metric("Issues logged", f"{len(dq):,}")
        st.dataframe(breakdown, use_container_width=True, hide_index=True)

    st.markdown("&nbsp;")
    st.markdown("**Sample fixes** — filter by rule")
    rules = ["(all)"] + sorted(dq["rule"].unique().tolist())
    f_rule = st.selectbox("Rule", rules, key="cln_rule")
    view = dq if f_rule == "(all)" else dq[dq["rule"] == f_rule]
    st.dataframe(
        view[["sheet", "row_seq", "column", "rule", "severity",
              "raw_value", "repaired_value", "note"]]
            .sort_values(["rule", "sheet", "row_seq"]),
        use_container_width=True, hide_index=True, height=420,
    )


def _tab_ai_mapping(mp: pd.DataFrame):
    st.markdown(
        "**Column-name decisions made by the LLM column mapper** "
        "(or the synonym matcher fallback when the foundation-model endpoint "
        "isn't reachable). Each carries a confidence score and rationale."
    )
    if mp.empty:
        st.info("No mappings recorded.")
        return

    visible = mp[mp["role"] != "ignored"].copy()
    src_count = visible.groupby("source").size().reset_index(name="n")
    quarantine = visible[visible["confidence"] < 0.5].shape[0]
    primary_source = src_count.sort_values("n", ascending=False).iloc[0]["source"] \
        if not src_count.empty else "—"

    s1, s2, s3 = st.columns(3)
    s1.metric("Total decisions", f"{len(visible):,}")
    s2.metric("Mapper", primary_source,
              help="`llm` = foundation model fired; `mock_synonyms` = deterministic fallback")
    s3.metric("Flagged for review", f"{quarantine}",
              delta="confidence < 0.5", delta_color="off")

    st.markdown("&nbsp;")
    st.caption("Showing only high-confidence mappings (confidence ≥ 0.90).")
    view = visible[visible["confidence"] >= 0.9].copy()
    view = view.sort_values("confidence")
    view["confidence"] = view["confidence"].round(2)
    st.dataframe(
        view[["sheet", "raw_label", "role", "canonical",
              "confidence", "source", "rationale"]],
        use_container_width=True, hide_index=True, height=460,
    )


def _tab_analytics(obs: pd.DataFrame):
    st.markdown(
        "**Compliance and trend views over the cleaned data from this run.**"
    )
    if obs.empty:
        st.info("No observations available.")
        return

    # Pass rate by batch
    st.markdown("#### Pass rate by batch")
    batch_stats = (
        obs.dropna(subset=["pass"])
           .groupby(["batch_no"])
           .agg(observations=("pass", "size"),
                passing=("pass", lambda s: int((s == True).sum())),
                failing=("pass", lambda s: int((s == False).sum())))
           .reset_index()
    )
    batch_stats["pass_rate_pct"] = (
        100.0 * batch_stats["passing"] /
        (batch_stats["passing"] + batch_stats["failing"]).replace(0, 1)
    ).round(1)
    batch_stats = batch_stats.sort_values("pass_rate_pct").head(20)
    if not batch_stats.empty:
        ca, cb = st.columns([2, 3])
        with ca:
            st.dataframe(batch_stats, use_container_width=True,
                         hide_index=True, height=380)
        with cb:
            st.bar_chart(batch_stats.set_index("batch_no")["pass_rate_pct"],
                         color=PRIMARY, height=380)

    st.markdown("&nbsp;")

    # Spec violations by analyte
    st.markdown("#### Spec violations by analyte")
    failing = obs[obs["pass"] == False]
    if not failing.empty:
        violations = (
            failing.groupby("analyte_canonical")
                   .agg(violations=("pass", "size"),
                        avg_value=("value", "mean"))
                   .reset_index()
                   .sort_values("violations", ascending=False)
        )
        violations["avg_value"] = violations["avg_value"].round(4)
        cv1, cv2 = st.columns([3, 2])
        with cv1:
            st.bar_chart(violations.set_index("analyte_canonical")["violations"],
                         color=ACCENT_RED, height=320)
        with cv2:
            st.dataframe(violations, use_container_width=True,
                         hide_index=True, height=320)
    else:
        st.success("Zero spec violations in this run.")

    st.markdown("&nbsp;")

    # Impurity trend
    st.markdown("#### Impurity trend (avg value over sample date)")
    if "sample_date" in obs.columns:
        with_dates = obs.dropna(subset=["sample_date", "value"]).copy()
        with_dates["sample_date"] = pd.to_datetime(with_dates["sample_date"])
        if not with_dates.empty:
            pivot = with_dates.pivot_table(
                index="sample_date", columns="analyte_canonical",
                values="value", aggfunc="mean",
            )
            st.line_chart(pivot, height=380)


def _list_volume_xlsx(vol_dir: str) -> list[dict]:
    """List xlsx files in a UC volume subdirectory via WorkspaceClient."""
    w = _ws_client()
    if w is None:
        return []
    out: list[dict] = []
    try:
        for f in w.files.list_directory_contents(vol_dir):
            if f.path.endswith(".xlsx"):
                out.append({
                    "name": Path(f.path).name,
                    "path": f.path,
                    "size": getattr(f, "file_size", 0) or 0,
                })
    except Exception as e:
        # Empty / not-yet-created folder is normal — return empty list
        print(f"  (volume listing skipped for {vol_dir}: {e})")
    return out


def _download_volume_bytes(remote_path: str) -> bytes:
    w = _ws_client()
    if w is None:
        return b""
    resp = w.files.download(remote_path)
    return resp.contents.read()


def _tab_outputs_session(cleaned_dir: str, tidy_dir: str):
    st.markdown(
        "**Both cleaned views are saved to this session's output folder.** "
        "Same-format mirrors the input shape; tidy is the long-form analytics view."
    )
    sub_same, sub_tidy = st.tabs([
        "Same-format (mirrors input)",
        "Tidy long-form (analytics)",
    ])

    with sub_same:
        st.caption(f"Folder: `{cleaned_dir}`")
        files = _list_volume_xlsx(cleaned_dir)
        if not files:
            st.info("No same-format outputs found.")
        else:
            choice = st.selectbox("File", [f["name"] for f in files],
                                  key="same_pick")
            sel = next(f for f in files if f["name"] == choice)
            data = _download_volume_bytes(sel["path"])
            st.download_button(
                "Download", data=data, file_name=sel["name"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key="dl_same",
                type="primary",
            )
            tmp = Path("/tmp/qde_preview") / sel["name"]
            tmp.parent.mkdir(exist_ok=True)
            tmp.write_bytes(data)
            try:
                sheets = _list_sheets(tmp)
                sheet = st.selectbox("Sheet", sheets, key="same_sheet_pick")
                _render_xlsx_full(tmp, sheet, height=620)
            except Exception as e:
                st.warning(f"Preview unavailable: {e}")

    with sub_tidy:
        st.caption(f"Folder: `{tidy_dir}`")
        files = _list_volume_xlsx(tidy_dir)
        if not files:
            st.info("No tidy outputs found.")
        else:
            choice = st.selectbox("File", [f["name"] for f in files],
                                  key="tidy_pick")
            sel = next(f for f in files if f["name"] == choice)
            data = _download_volume_bytes(sel["path"])
            st.download_button(
                "Download", data=data, file_name=sel["name"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key="dl_tidy",
                type="primary",
            )
            tmp = Path("/tmp/qde_preview") / sel["name"]
            tmp.parent.mkdir(exist_ok=True)
            tmp.write_bytes(data)
            try:
                sheets = _list_sheets(tmp)
                sheet = st.selectbox("Sheet", sheets, key="tidy_sheet_pick")
                _render_xlsx_full(tmp, sheet, height=620)
            except Exception as e:
                st.warning(f"Preview unavailable: {e}")


def render_home():
    stage = st.session_state.stage
    if stage == "empty":
        _render_home_empty()
    elif stage == "staged":
        _render_home_staged()
    elif stage == "running":
        _render_home_running()
    elif stage == "results":
        _render_home_results()
    else:
        st.session_state.stage = "empty"
        st.rerun()


# ===========================================================================
# History view
# ===========================================================================

def render_history():
    st.title("History")
    st.markdown(
        "<div style='opacity:0.75;font-size:1.0rem;margin-bottom:1rem;'>"
        "Every interactive run, newest first. Click <strong>Open</strong> to re-load "
        "that run's results." "</div>",
        unsafe_allow_html=True,
    )

    runs = _list_runs()
    if not runs:
        st.info("No runs yet. Head to Home and process a workbook.")
        return

    # Build a display dataframe
    df = pd.DataFrame(runs)
    df["started"] = pd.to_datetime(df["started_at"]).dt.strftime("%Y-%m-%d %H:%M UTC")
    show_cols = [
        "run_id", "file_name", "started", "duration_s",
        "n_observations", "n_dq_issues", "n_violations", "pass_rate_pct",
    ]
    display = df[show_cols].rename(columns={
        "run_id": "Run ID", "file_name": "File", "started": "When",
        "duration_s": "Duration (s)",
        "n_observations": "Observations", "n_dq_issues": "DQ fixes",
        "n_violations": "Violations", "pass_rate_pct": "Pass %",
    })

    # KPI strip across all runs
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Runs", f"{len(df):,}")
    k2.metric("Total observations", f"{df['n_observations'].sum():,}")
    k3.metric("Total DQ fixes", f"{df['n_dq_issues'].sum():,}")
    k4.metric("Avg pass rate", f"{df['pass_rate_pct'].mean():.1f}%")

    st.markdown("&nbsp;")

    # Use selection-enabled dataframe
    selection = st.dataframe(
        display, use_container_width=True, hide_index=True, height=420,
        on_select="rerun", selection_mode="single-row",
    )
    sel_rows = selection.selection.rows if selection and selection.selection else []
    if sel_rows:
        sel_run = display.iloc[sel_rows[0]]["Run ID"]
        c1, c2 = st.columns([3, 1])
        with c1:
            st.info(f"Selected run **{sel_run}**.")
        with c2:
            if st.button("Open this run →", type="primary",
                         use_container_width=True, key="open_hist"):
                st.session_state.current_run_id = sel_run
                st.session_state.view = "home"
                st.session_state.stage = "results"
                st.rerun()


# ===========================================================================
# Dashboard view
# ===========================================================================

def render_dashboard():
    st.title("Dashboard")
    st.markdown(
        "<div style='opacity:0.75;font-size:1.0rem;margin-bottom:1rem;'>"
        "Comprehensive analytics — both this app's interactive runs <em>and</em> "
        "the batch pipeline's bronze/silver/gold layer."
        "</div>",
        unsafe_allow_html=True,
    )

    # ── Section A: Interactive runs ─────────────────────────────────────────
    runs = _list_runs()
    st.markdown("## App runs (interactive)")
    if not runs:
        st.info("No interactive runs yet.")
    else:
        df = pd.DataFrame(runs)
        df["started"] = pd.to_datetime(df["started_at"])

        a1, a2, a3, a4, a5 = st.columns(5)
        a1.metric("Runs", f"{len(df):,}")
        a2.metric("Files cleaned", f"{df['file_name'].count():,}")
        a3.metric("Total obs", f"{df['n_observations'].sum():,}")
        a4.metric("Total fixes", f"{df['n_dq_issues'].sum():,}")
        a5.metric("Avg pass %", f"{df['pass_rate_pct'].mean():.1f}%")

        st.markdown("&nbsp;")

        # Charts
        c1, c2 = st.columns(2, gap="large")
        with c1:
            st.markdown("**Runs over time**")
            by_day = (df.set_index("started")
                        .resample("D").size()
                        .rename("runs"))
            if not by_day.empty:
                st.bar_chart(by_day, color=PRIMARY, height=260)

        with c2:
            st.markdown("**Avg pass rate per run**")
            ser = df.set_index("started")["pass_rate_pct"].sort_index()
            if not ser.empty:
                st.line_chart(ser, color=ACCENT_GREEN, height=260)

    st.markdown("---")

    # ── Section B: Batch pipeline ───────────────────────────────────────────
    st.markdown("## Batch pipeline (bronze / silver / gold)")
    st.caption(
        "Aggregates from the deployed pipeline tables — these are populated by "
        "the scheduled medallion job (`./deploy.sh` runs)."
    )

    try:
        batch_summary = run_query(f"""
            SELECT
              (SELECT COUNT(*) FROM {CATALOG}.bronze.raw_workbooks)            AS files,
              (SELECT COUNT(*) FROM {CATALOG}.silver.observations_long)        AS observations,
              (SELECT COUNT(*) FROM {CATALOG}.silver.dq_issues)                AS fixes,
              (SELECT COUNT(*) FROM {CATALOG}.silver.column_mapping_log)       AS mappings,
              (SELECT COUNT(*) FROM {CATALOG}.gold.fact_observation
                 WHERE pass = false)                                          AS violations,
              (SELECT ROUND(AVG(CASE WHEN pass=true THEN 100.0
                                     WHEN pass=false THEN 0.0 END), 1)
                 FROM {CATALOG}.gold.fact_observation)                        AS pass_rate
        """)
    except Exception as e:
        batch_summary = pd.DataFrame()

    if batch_summary.empty:
        st.info(
            "Batch pipeline tables not reachable. Make sure the bundle is deployed "
            "and the pipeline job has been run at least once."
        )
        return

    row = batch_summary.iloc[0]
    b1, b2, b3, b4, b5, b6 = st.columns(6)
    b1.metric("Files ingested", f"{int(row.files):,}")
    b2.metric("Observations", f"{int(row.observations):,}")
    b3.metric("DQ fixes", f"{int(row.fixes):,}")
    b4.metric("AI mappings", f"{int(row.mappings):,}")
    b5.metric("Spec violations", f"{int(row.violations):,}")
    b6.metric("Pass rate", f"{row.pass_rate}%")

    st.markdown("&nbsp;")

    # Top spec violators + DQ rule breakdown
    cb1, cb2 = st.columns(2, gap="large")
    with cb1:
        st.markdown("**Top spec violators (gold)**")
        try:
            tv = run_query(f"""
                SELECT analyte_canonical, SUM(violation_count) AS violations
                FROM {CATALOG}.gold.mv_spec_violations
                GROUP BY analyte_canonical
                ORDER BY violations DESC LIMIT 10
            """)
            if not tv.empty:
                st.bar_chart(tv.set_index("analyte_canonical")["violations"],
                             color=ACCENT_RED, height=320)
        except Exception:
            st.warning("Couldn't load spec violations.")

    with cb2:
        st.markdown("**DQ rule breakdown (silver)**")
        try:
            dq_break = run_query(f"""
                SELECT rule, COUNT(*) AS issues
                FROM {CATALOG}.silver.dq_issues
                GROUP BY rule ORDER BY issues DESC
            """)
            if not dq_break.empty:
                st.bar_chart(dq_break.set_index("rule")["issues"],
                             color=ACCENT_AMBER, height=320)
        except Exception:
            st.warning("Couldn't load DQ rule breakdown.")

    # Impurity trend
    st.markdown("&nbsp;")
    st.markdown("**Impurity trend (avg value, gold.mv_impurity_trend)**")
    try:
        trend = run_query(f"""
            SELECT analyte_canonical, sample_date, avg_value
            FROM {CATALOG}.gold.mv_impurity_trend
            ORDER BY sample_date
        """)
        if not trend.empty:
            pivot = trend.pivot_table(
                index="sample_date", columns="analyte_canonical",
                values="avg_value", aggfunc="mean",
            )
            st.line_chart(pivot, height=380)
    except Exception:
        st.warning("Couldn't load trend.")


# ===========================================================================
# Main
# ===========================================================================

_sidebar()

view = st.session_state.view
if view == "home":
    render_home()
elif view == "history":
    render_history()
elif view == "dashboard":
    render_dashboard()
else:
    st.session_state.view = "home"
    st.rerun()


# Footer
st.markdown(
    f"""
    <div class='footer-note'>
       <strong>Quality Team Intelligence</strong> · catalog <code>{CATALOG}</code> ·
       runs cached at <code>{RUNS_DIR}</code> (per-session) ·
       <a href='https://github.com/AbhinavJFT/primeinsurance-poc' target='_blank'>repo</a>
    </div>
    """,
    unsafe_allow_html=True,
)
