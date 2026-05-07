"""Synthetic Quality team workbook generator.

Produces multi-sheet pharmaceutical .xlsx files shaped like the real Quality
team workbooks shown in the project screenshots: HPLC batch testing data
with merged-cell headers, RT/RRT/Specification metadata bands above the
data, and realistic dirt (typos, blanks, wrong-format dates/times, NULL/?
sentinels, occasional out-of-range values).

Mirrors the pattern from dpdp/generate_salesforce_data.py — deterministic
seed, no external I/O, callable as both a library and CLI tool.

Usage:
    python generate_quality_data.py --out data/mock_sharepoint/input/ --seed 43
    python generate_quality_data.py --workbook api --sample             # dry-run preview
"""

from __future__ import annotations

import argparse
import random
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


DEFAULT_SEED = 43
GENERATOR_DATE = date(2026, 3, 14)


# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
GREY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BLUE_FONT = Font(color="2F5496", bold=True)
PURPLE_FONT = Font(color="7030A0", bold=True)
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
DATA_ALIGN = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Sheet specifications
# ---------------------------------------------------------------------------

@dataclass
class ImpuritySpec:
    label: str           # what appears in the header (often inconsistent)
    rt: float
    rrt: float
    unit: str = "%"
    sub_unit: str = "v/v"
    bound: str = "Max"   # "Max" or "Min"
    spec: float = 0.10
    internal_spec: float = 0.08
    typical: float = 0.005    # mean of generated values
    spread: float = 0.003


@dataclass
class SheetSpec:
    sheet_name: str
    process_label: str           # e.g. "Inprocess REACTION MASS"
    stage_value: str             # e.g. "Reprocess"
    sample_form: str             # e.g. "Solid"
    appearance: str              # e.g. "White Powder"
    appearance_solution_desc: str  # the long text in the unit row
    instrument: str              # e.g. "HPLC-02/KF-12"
    impurities: list[ImpuritySpec]
    starting_serial: int
    starting_batch_no: int
    n_rows: int = 50
    batch_prefix: str = "#"


@dataclass
class WorkbookSpec:
    filename: str
    sheets: list[SheetSpec]
    misc: bool = True            # whether to append a MISC sheet


# Real-screenshot-shaped sheet templates ------------------------------------

def _api_sheets() -> list[SheetSpec]:
    common_imps_a = [
        ImpuritySpec("Impurity-1",          0.70, 0.55, bound="Max", spec=0.10, internal_spec=0.08, typical=0.001, spread=0.001),
        ImpuritySpec("2-chloro benzamide",  1.06, 0.83, bound="Max", spec=0.10, internal_spec=0.08, typical=0.007, spread=0.003),
        ImpuritySpec("Impurity-1",          1.12, 0.88, bound="Max", spec=0.10, internal_spec=0.08, typical=0.001, spread=0.001),
        ImpuritySpec("TFMBA",               1.27, 1.00, bound="Min", spec=99.00, internal_spec=99.00, typical=99.85, spread=0.05),
        ImpuritySpec("Impurity",            1.60, 1.26, bound="Max", spec=0.10, internal_spec=0.08, typical=0.001, spread=0.001),
        ImpuritySpec("5-chloro TFMBA",      1.80, 1.42, bound="Max", spec=0.20, internal_spec=0.16, typical=0.025, spread=0.012),
        ImpuritySpec("Impurity-1",          1.88, 1.48, bound="Max", spec=0.10, internal_spec=0.08, typical=0.001, spread=0.001),
        ImpuritySpec("Impurity-2",          2.02, 1.59, bound="Max", spec=0.10, internal_spec=0.08, typical=0.001, spread=0.001),
        ImpuritySpec("TFMBA Acid",          2.16, 1.70, bound="Max", spec=0.40, internal_spec=0.32, typical=0.005, spread=0.003),
    ]
    common_imps_b = [
        ImpuritySpec("Impurity-1",          0.65, 0.51, bound="Max", spec=0.15, internal_spec=0.12, typical=0.002, spread=0.002),
        ImpuritySpec("2-chloro benzamide",  0.98, 0.78, bound="Max", spec=0.15, internal_spec=0.12, typical=0.009, spread=0.004),
        ImpuritySpec("TFMBA",               1.27, 1.00, bound="Min", spec=98.50, internal_spec=98.50, typical=99.20, spread=0.20),
        ImpuritySpec("5-chloro TFMBA",      1.85, 1.46, bound="Max", spec=0.25, internal_spec=0.20, typical=0.040, spread=0.020),
        ImpuritySpec("TFMBA Acid",          2.20, 1.73, bound="Max", spec=0.50, internal_spec=0.40, typical=0.012, spread=0.006),
    ]
    return [
        SheetSpec(
            sheet_name="1. R-2127 & R-2126",
            process_label="Inprocess REACTION MASS",
            stage_value="Reaction Mass",
            sample_form="Liquid",
            appearance="Clear Brown Liquid",
            appearance_solution_desc="Clear colorless soultion without any black particle and settling matter",
            instrument="HPLC-02/KF-12",
            impurities=common_imps_b,
            starting_serial=101,
            starting_batch_no=2127,
            n_rows=42,
            batch_prefix="R-",
        ),
        SheetSpec(
            sheet_name="2. V-2424",
            process_label="Inprocess VESSEL TRANSFER",
            stage_value="Transfer",
            sample_form="Liquid",
            appearance="Pale Yellow Liquid",
            appearance_solution_desc="Clear colorless soultion without any black particle and settling matter",
            instrument="HPLC-04/KF-08",
            impurities=common_imps_b,
            starting_serial=201,
            starting_batch_no=2424,
            n_rows=38,
            batch_prefix="V-",
        ),
        SheetSpec(
            sheet_name="3. ANF-2302 AFTER WASH",
            process_label="Inprocess ANF AFTER WASH",
            stage_value="After Wash",
            sample_form="Wet Cake",
            appearance="Off-white wet cake",
            appearance_solution_desc="Clear colorless soultion without any black particle and settling matter",
            instrument="HPLC-02/KF-12",
            impurities=common_imps_a,
            starting_serial=301,
            starting_batch_no=2302,
            n_rows=44,
            batch_prefix="ANF-",
        ),
        SheetSpec(
            sheet_name="4. D-2301",
            process_label="Inprocess REACTION MASS",
            stage_value="Reprocess",
            sample_form="Solid",
            appearance="White Powder",
            appearance_solution_desc="Clear colorless soultion without any black particle and settling matter",
            instrument="HPLC-02/KF-12",
            impurities=common_imps_a,
            starting_serial=522,
            starting_batch_no=710,
            n_rows=52,
            batch_prefix="#",
        ),
        SheetSpec(
            sheet_name="4.D-2302",
            process_label="Inprocess REACTION MASS",
            stage_value="Reprocess",
            sample_form="Solid",
            appearance="White Powder",
            appearance_solution_desc="Clear colorless soultion without any black particle and settling matter",
            instrument="HPLC-02/KF-12",
            impurities=common_imps_a,
            starting_serial=580,
            starting_batch_no=820,
            n_rows=40,
            batch_prefix="#",
        ),
        SheetSpec(
            sheet_name="T-2203-TOP",
            process_label="Inprocess CRYSTALLIZATION",
            stage_value="Crystallization (Top Cut)",
            sample_form="Solid",
            appearance="White Crystalline",
            appearance_solution_desc="Clear colorless soultion without any black particle and settling matter",
            instrument="HPLC-05/KF-04",
            impurities=common_imps_a[:6],
            starting_serial=901,
            starting_batch_no=2203,
            n_rows=36,
            batch_prefix="T-",
        ),
    ]


def _ksm_sheets() -> list[SheetSpec]:
    imps = [
        ImpuritySpec("Imp-A",       0.60, 0.48, bound="Max", spec=0.20, internal_spec=0.15, typical=0.04, spread=0.02),
        ImpuritySpec("Imp-B",       0.92, 0.74, bound="Max", spec=0.20, internal_spec=0.15, typical=0.03, spread=0.02),
        ImpuritySpec("KSM Assay",   1.20, 1.00, bound="Min", spec=98.00, internal_spec=98.50, typical=99.40, spread=0.30),
        ImpuritySpec("Imp-D",       1.55, 1.29, bound="Max", spec=0.30, internal_spec=0.25, typical=0.08, spread=0.04),
        ImpuritySpec("Imp-E",       1.92, 1.60, bound="Max", spec=0.50, internal_spec=0.40, typical=0.15, spread=0.07),
    ]
    return [
        SheetSpec(
            sheet_name="1. KSM-A101",
            process_label="Inprocess KSM PURIFICATION",
            stage_value="Purification",
            sample_form="Solid",
            appearance="Off-white powder",
            appearance_solution_desc="Clear pale yellow solution, free of particulate matter",
            instrument="HPLC-01/KF-02",
            impurities=imps,
            starting_serial=1, starting_batch_no=101, n_rows=34, batch_prefix="KSM-A",
        ),
        SheetSpec(
            sheet_name="2. KSM-B204 STAGE 2",
            process_label="Inprocess KSM STAGE 2",
            stage_value="Stage 2",
            sample_form="Solid",
            appearance="White powder",
            appearance_solution_desc="Clear colourless solution, free of particulate matter",
            instrument="HPLC-01/KF-02",
            impurities=imps,
            starting_serial=1, starting_batch_no=204, n_rows=30, batch_prefix="KSM-B",
        ),
    ]


def _intermediate_sheets() -> list[SheetSpec]:
    imps = [
        ImpuritySpec("Impurity-1",       0.72, 0.58, bound="Max", spec=0.15, internal_spec=0.12, typical=0.02, spread=0.012),
        ImpuritySpec("Impurity-2",       1.05, 0.84, bound="Max", spec=0.15, internal_spec=0.12, typical=0.025, spread=0.014),
        ImpuritySpec("Main Compound",    1.25, 1.00, bound="Min", spec=98.50, internal_spec=99.00, typical=99.50, spread=0.20),
        ImpuritySpec("Impurity-3",       1.66, 1.33, bound="Max", spec=0.20, internal_spec=0.16, typical=0.04, spread=0.02),
        ImpuritySpec("Unknown-RRT-1.7",  2.12, 1.70, bound="Max", spec=0.10, internal_spec=0.08, typical=0.008, spread=0.005),
    ]
    return [
        SheetSpec("1. INT-301", "Inprocess INTERMEDIATE", "Reaction", "Solid", "Light yellow powder",
                  "Clear pale yellow solution, free of particulate matter", "HPLC-03/KF-06",
                  imps, 1, 301, 38, "INT-"),
        SheetSpec("2. INT-302 ALT", "Inprocess INTERMEDIATE ALT ROUTE", "Reaction (alt)", "Solid",
                  "Off-white powder", "Clear colourless solution, free of particulate matter",
                  "HPLC-03/KF-06", imps, 1, 302, 32, "INT-"),
        SheetSpec("3. INT-303 BLEND", "Inprocess BLEND", "Blend", "Solid", "White powder",
                  "Clear colourless solution, free of particulate matter", "HPLC-03/KF-06",
                  imps, 1, 303, 28, "INT-"),
    ]


def workbook_specs() -> dict[str, WorkbookSpec]:
    return {
        "api": WorkbookSpec(
            filename="QualityBook_2026Q1_API.xlsx",
            sheets=_api_sheets(),
            misc=True,
        ),
        "ksm": WorkbookSpec(
            filename="QualityBook_2026Q1_KSM.xlsx",
            sheets=_ksm_sheets(),
            misc=True,
        ),
        "intermediates": WorkbookSpec(
            filename="QualityBook_2026Q1_Intermediates.xlsx",
            sheets=_intermediate_sheets(),
            misc=False,
        ),
    }


# ---------------------------------------------------------------------------
# Dirt helpers
# ---------------------------------------------------------------------------

def _dirty_ok(rng: random.Random) -> str:
    return rng.choices(
        ["OK", "Ok", "ok", " OK ", "OK ", "NULL", "?", "", "OK"],
        weights=[55, 8, 5, 4, 4, 3, 2, 4, 15],
    )[0]


def _dirty_appearance(rng: random.Random, base: str) -> str:
    return rng.choices(
        [base, base.lower(), f" {base} ", base.upper(), "", "NULL", base],
        weights=[60, 8, 6, 4, 5, 2, 15],
    )[0]


def _dirty_time(rng: random.Random, t: time) -> object:
    """Return a real datetime.time, or sometimes a malformed string."""
    if rng.random() < 0.04:
        return rng.choice(["27:00.0", f"{t.hour}.{t.minute:02d}", "??:??", ""])
    return t


def _dirty_date(rng: random.Random, d: date) -> object:
    if rng.random() < 0.03:
        return rng.choice([d.strftime("%d/%m/%Y"), d.strftime("%b %d %Y"), ""])
    return d


def _dirty_value(rng: random.Random, v: float, dirt_p: float = 0.04) -> object:
    if rng.random() < dirt_p:
        return rng.choice([f" {v:.4f}", f"{v:.4f} ", "NULL", "?", "", f"{-v:.4f}"])
    return round(v, 4)


def _next_serial(start: int, i: int) -> int:
    return start + i


def _shift_date(rng: random.Random, base: date, idx: int) -> date:
    return base + timedelta(days=idx + rng.randint(0, 1))


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

# Column layout for the standard quality sheet:
#   A = Serial number (header label "REACTION MASS" in row 4)
#   B = Date of Sampling
#   C = Time of sampling
#   D = Time of Reporting
#   E = Batch No.
#   F = Instrument ID
#   G = Stage
#   H = Sample Characteristic
#   I = (spacer column, often blank in source)
#   J = Appearance
#   K = Appearance of 30% solution in methanol
#   L..    = impurity columns

HEADER_ROW_PROCESS = 3       # "Inprocess REACTION MASS" / "Analysis"
HEADER_ROW_LABEL = 4         # field labels
HEADER_ROW_RT = 5
HEADER_ROW_RRT = 6
HEADER_ROW_UNIT = 7
HEADER_ROW_SUBUNIT = 8
HEADER_ROW_SPEC_BOUND = 9
HEADER_ROW_SPEC_VALUE = 10
HEADER_ROW_INTERNAL_SPEC = 11
DATA_START_ROW = 13          # leave row 12 blank like the screenshots


def _write_quality_sheet(ws: Worksheet, spec: SheetSpec, rng: random.Random) -> None:
    n_imp = len(spec.impurities)
    first_imp_col = 12   # column L
    last_imp_col = first_imp_col + n_imp - 1
    last_col_letter = get_column_letter(last_imp_col)

    # Row 3: process label (left side) + Analysis (right side)
    ws.cell(row=HEADER_ROW_PROCESS, column=1, value=spec.process_label).font = Font(bold=True)
    ws.merge_cells(start_row=HEADER_ROW_PROCESS, start_column=1,
                   end_row=HEADER_ROW_PROCESS, end_column=9)
    analysis_cell = ws.cell(row=HEADER_ROW_PROCESS, column=10, value="Analysis")
    analysis_cell.font = Font(bold=True)
    ws.merge_cells(start_row=HEADER_ROW_PROCESS, start_column=10,
                   end_row=HEADER_ROW_PROCESS, end_column=last_imp_col)

    # Row 4: field labels
    labels_left = {
        1: "REACTION MASS",
        2: "Date of Sampling",
        3: "Time of sampling",
        4: "Time of Reporting",
        5: "Batch No.",
        6: "Instrument ID",
        7: "Stage",
        8: "Sample Characteristic",
        10: "Appearance",
        11: "Appearance of 30% solution in methanol (after 12 hrs settling)",
    }
    for col, label in labels_left.items():
        c = ws.cell(row=HEADER_ROW_LABEL, column=col, value=label)
        c.font = Font(bold=True)
        c.alignment = HEADER_ALIGN
    for j, imp in enumerate(spec.impurities):
        c = ws.cell(row=HEADER_ROW_LABEL, column=first_imp_col + j, value=imp.label)
        c.font = BLUE_FONT
        c.alignment = HEADER_ALIGN

    # Row 5: RT
    ws.cell(row=HEADER_ROW_RT, column=1, value="RT").font = Font(bold=True)
    for j, imp in enumerate(spec.impurities):
        ws.cell(row=HEADER_ROW_RT, column=first_imp_col + j, value=imp.rt)

    # Row 6: RRT
    ws.cell(row=HEADER_ROW_RRT, column=1, value="RRT").font = Font(bold=True)
    for j, imp in enumerate(spec.impurities):
        ws.cell(row=HEADER_ROW_RRT, column=first_imp_col + j, value=imp.rrt)

    # Row 7: Unit
    ws.cell(row=HEADER_ROW_UNIT, column=1, value="Unit").font = Font(bold=True)
    desc_cell = ws.cell(row=HEADER_ROW_UNIT, column=11, value=spec.appearance_solution_desc)
    desc_cell.alignment = HEADER_ALIGN
    for j, imp in enumerate(spec.impurities):
        ws.cell(row=HEADER_ROW_UNIT, column=first_imp_col + j, value=imp.unit)

    # Row 8: sub-unit (v/v)
    for j, imp in enumerate(spec.impurities):
        ws.cell(row=HEADER_ROW_SUBUNIT, column=first_imp_col + j, value=imp.sub_unit)

    # Row 9: Specification + bound (Max/Min)
    ws.cell(row=HEADER_ROW_SPEC_BOUND, column=1, value="Specification").font = Font(bold=True)
    ws.cell(row=HEADER_ROW_SPEC_BOUND, column=11, value="0.0000")
    for j, imp in enumerate(spec.impurities):
        c = ws.cell(row=HEADER_ROW_SPEC_BOUND, column=first_imp_col + j, value=imp.bound)
        c.font = BLUE_FONT

    # Row 10: spec values
    for j, imp in enumerate(spec.impurities):
        c = ws.cell(row=HEADER_ROW_SPEC_VALUE, column=first_imp_col + j, value=imp.spec)
        c.font = BLUE_FONT

    # Row 11: Internal SRF spec (highlighted greenish in screenshots)
    isrf = ws.cell(row=HEADER_ROW_INTERNAL_SPEC, column=1, value="Internal SRF Specification")
    isrf.font = Font(bold=True)
    isrf.fill = LIGHT_GREEN_FILL
    for col in range(2, last_imp_col + 1):
        ws.cell(row=HEADER_ROW_INTERNAL_SPEC, column=col).fill = LIGHT_GREEN_FILL
    for j, imp in enumerate(spec.impurities):
        c = ws.cell(row=HEADER_ROW_INTERNAL_SPEC, column=first_imp_col + j, value=imp.internal_spec)
        c.font = PURPLE_FONT

    # Data rows
    base_date = GENERATOR_DATE
    base_sample_time = time(hour=8, minute=0)
    for i in range(spec.n_rows):
        row = DATA_START_ROW + i
        serial = _next_serial(spec.starting_serial, i)
        d = _shift_date(rng, base_date, i)
        sample_t = (datetime.combine(date.today(), base_sample_time)
                    + timedelta(hours=rng.randint(0, 14), minutes=rng.choice([0, 15, 20, 30, 35, 40, 50]))).time()
        report_t = (datetime.combine(date.today(), sample_t)
                    + timedelta(hours=rng.randint(2, 6), minutes=rng.randint(0, 59))).time()

        ws.cell(row=row, column=1, value=serial)
        ws.cell(row=row, column=2, value=_dirty_date(rng, d))
        ws.cell(row=row, column=3, value=_dirty_time(rng, sample_t))
        ws.cell(row=row, column=4, value=_dirty_time(rng, report_t))
        ws.cell(row=row, column=5,
                value=f"{spec.batch_prefix}{spec.starting_batch_no + i}")
        ws.cell(row=row, column=6, value=spec.instrument)
        ws.cell(row=row, column=7, value=spec.stage_value)
        ws.cell(row=row, column=8, value=spec.sample_form)
        # column 9 (I) intentionally blank — matches screenshots
        ws.cell(row=row, column=10, value=_dirty_appearance(rng, spec.appearance))
        ws.cell(row=row, column=11, value=_dirty_ok(rng))

        for j, imp in enumerate(spec.impurities):
            base = rng.gauss(imp.typical, imp.spread)
            # Occasionally spike a value out-of-spec so the dashboard has real failures.
            if rng.random() < 0.04:
                if imp.bound == "Max":
                    base = imp.spec * rng.uniform(1.05, 1.8)
                else:  # "Min" spec — fall slightly below
                    base = imp.spec * rng.uniform(0.985, 0.998)
            elif imp.bound == "Min":
                base = max(min(base, 100.0), imp.spec - rng.uniform(0, 0.5))
            else:
                base = abs(base)
            cell = ws.cell(row=row, column=first_imp_col + j,
                           value=_dirty_value(rng, base))
            # green highlight for in-spec rows on the "main" assay column (typical of screenshots)
            if imp.bound == "Min" and isinstance(cell.value, (int, float)) and cell.value >= imp.spec:
                cell.fill = LIGHT_GREEN_FILL

    # Column widths (cosmetic — helps the demo look real when opened in Excel)
    ws.column_dimensions["A"].width = 6
    for col_letter, w in {"B": 12, "C": 11, "D": 11, "E": 11, "F": 16, "G": 18,
                          "H": 18, "I": 4, "J": 22, "K": 28}.items():
        ws.column_dimensions[col_letter].width = w
    for j in range(n_imp):
        ws.column_dimensions[get_column_letter(first_imp_col + j)].width = 13


def _write_misc_sheet(ws: Worksheet, rng: random.Random) -> None:
    """A free-form sheet with a different structure to stress the cleaner."""
    ws.cell(row=1, column=1, value="Miscellaneous QA log").font = Font(bold=True, size=14)
    headers = ["S.No", "Date", "Description", "Test ID", "Result", "Reviewer", "Notes"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = GREY_FILL

    descriptions = [
        "Calibration check on HPLC-02",
        "Out-of-trend impurity in batch",
        "Solvent system rinse audit",
        "Column pressure spike investigation",
        "Reagent expiry verification",
        "Daily system suitability test",
        "Re-injection of suspect sample",
        "Method transfer dry run",
    ]
    reviewers = ["S. Iyer", "N. Gupta", "A. Khan", "P. Rao", "R. Mehra"]
    for i in range(28):
        row = 4 + i
        d = GENERATOR_DATE + timedelta(days=i)
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=_dirty_date(rng, d))
        ws.cell(row=row, column=3, value=rng.choice(descriptions))
        ws.cell(row=row, column=4, value=f"QA-{rng.randint(1000, 9999)}")
        ws.cell(row=row, column=5, value=rng.choices(["Pass", "Fail", "Review", "OK", "NULL", ""],
                                                     weights=[55, 8, 12, 15, 4, 6])[0])
        ws.cell(row=row, column=6, value=rng.choice(reviewers))
        ws.cell(row=row, column=7, value=rng.choice([
            "Within limits", "Repeat after recalibration", "Escalated to QC head",
            "No deviation", "", "Pending review", "Documented"]))

    for col_letter, w in {"A": 6, "B": 12, "C": 36, "D": 12, "E": 10, "F": 14, "G": 28}.items():
        ws.column_dimensions[col_letter].width = w


# ---------------------------------------------------------------------------
# Workbook assembly
# ---------------------------------------------------------------------------

def build_workbook(spec: WorkbookSpec, seed: int = DEFAULT_SEED) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    rng = random.Random(seed)
    for sheet_spec in spec.sheets:
        # spawn a per-sheet child rng so adding/removing sheets doesn't reshuffle others
        sheet_rng = random.Random(rng.randint(0, 2**31))
        ws = wb.create_sheet(title=sheet_spec.sheet_name)
        _write_quality_sheet(ws, sheet_spec, sheet_rng)
    if spec.misc:
        misc_rng = random.Random(rng.randint(0, 2**31))
        ws = wb.create_sheet(title="MISC")
        _write_misc_sheet(ws, misc_rng)
    return wb


def generate(out_dir: Path | str, seed: int = DEFAULT_SEED, only: list[str] | None = None) -> list[Path]:
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    specs = workbook_specs()
    target_keys = list(specs) if not only else [k for k in only if k in specs]
    for offset, key in enumerate(target_keys):
        spec = specs[key]
        wb = build_workbook(spec, seed=seed + offset)
        target = out / spec.filename
        wb.save(target)
        written.append(target)
    return written


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _main() -> None:
    parser = argparse.ArgumentParser(description="Generate synthetic Quality team workbooks.")
    parser.add_argument("--out", default="data/mock_sharepoint/input",
                        help="Output directory")
    parser.add_argument("--seed", type=int, default=DEFAULT_SEED)
    parser.add_argument("--workbook", choices=list(workbook_specs()), action="append",
                        help="Restrict to specific workbook key(s)")
    parser.add_argument("--sample", action="store_true",
                        help="Print a 5-row sample of the first sheet of the first workbook")
    args = parser.parse_args()

    if args.sample:
        spec = next(iter(workbook_specs().values())) if not args.workbook \
            else workbook_specs()[args.workbook[0]]
        first_sheet = spec.sheets[0]
        print(f"Workbook: {spec.filename}")
        print(f"  First sheet: {first_sheet.sheet_name}")
        print(f"  Process label: {first_sheet.process_label}")
        print(f"  Impurities ({len(first_sheet.impurities)}):")
        for imp in first_sheet.impurities:
            print(f"    - {imp.label:<24} RT={imp.rt}  RRT={imp.rrt}  "
                  f"{imp.bound} {imp.spec}  internal={imp.internal_spec}")
        print(f"  Data rows: {first_sheet.n_rows}")
        return

    written = generate(args.out, seed=args.seed, only=args.workbook)
    for p in written:
        print(f"wrote {p}  ({p.stat().st_size:,} bytes)")


if __name__ == "__main__":
    _main()
